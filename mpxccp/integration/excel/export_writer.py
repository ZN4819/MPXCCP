from __future__ import annotations

from copy import copy
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mpxccp.integration.excel.schema import (
    APPLICATION_ACCESS_COLUMNS,
    APPLICATION_BUSINESS_COLUMNS,
    APPLICATION_IMPORTANT_DATA_COLUMNS,
    APPLICATION_SECTIONS,
    APPLICATION_USER_COLUMNS,
    DEVICE_AUTH_COLUMNS,
    DEVICE_INTEGRITY_COLUMNS,
    DEVICE_OBJECT_COLUMNS,
    DEVICE_REMOTE_COLUMNS,
    NETWORK_AUTH_COLUMNS,
    NETWORK_BOUNDARY_COLUMNS,
    NETWORK_CONFIDENTIALITY_COLUMNS,
    NETWORK_INTEGRITY_COLUMNS,
    NETWORK_OBJECT_COLUMNS,
    PHYSICAL_AUTH_COLUMNS,
    PHYSICAL_INTEGRITY_COLUMNS,
    PHYSICAL_OBJECT_COLUMNS,
    SHEET_NAMES,
)
from mpxccp.integration.excel.workbook_styles import (
    apply_header_row,
    apply_table_style,
    set_column_widths,
)
from mpxccp.services.application_service import (
    ApplicationService,
)
from mpxccp.services.basic_info_service import BasicInfoService
from mpxccp.services.device_service import DeviceService
from mpxccp.services.network_service import NetworkService
from mpxccp.services.physical_service import PhysicalService
from mpxccp.services.quant_service import QuantService

MODULE_SHEETS = {
    "basic_info": SHEET_NAMES.basic_info,
    "physical": SHEET_NAMES.physical,
    "device": SHEET_NAMES.device,
    "network": SHEET_NAMES.network,
    "application": SHEET_NAMES.application,
}
ALL_DATA_SHEETS = tuple(MODULE_SHEETS.values())


def _data_unit_headers(
    prefix: str,
    *,
    include_channel: bool,
    include_mitigation: bool,
) -> list[str]:
    headers = [
        f"{prefix}判定",
        f"{prefix}算法",
        f"{prefix}机制说明",
        f"{prefix}图片",
        f"{prefix}技术合规",
        f"{prefix}产品合规",
        f"{prefix}产品",
        "D",
        "A",
        "K",
        "Ra",
        "Rk",
    ]
    if include_channel:
        headers.append(f"{prefix}关联信道")
    headers.extend([f"{prefix}密钥管理", f"{prefix}风险等级", f"{prefix}风险分析"])
    if include_mitigation:
        headers.extend([f"{prefix}缓释机制", f"{prefix}缓释说明", f"{prefix}缓释后风险等级"])
    headers.append(f"{prefix}整改建议")
    return headers


PHYSICAL_HEADERS = (
    "测评对象",
    "现场访谈记录",
    "物理地址",
    "门禁系统",
    "视频监控系统",
    "身份鉴别方式",
    "密码技术使用",
    "算法",
    "技术合规",
    "产品合规",
    "身份鉴别密码产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "专人值守",
    "进出登记",
    "专人陪同",
    "实时监控",
    "身份鉴别风险等级",
    "身份鉴别风险分析",
    "身份鉴别缓释机制",
    "身份鉴别缓释机制说明",
    "身份鉴别缓释后风险等级",
    "身份鉴别整改建议",
    "门禁记录完整性",
    "门禁算法",
    "门禁技术合规",
    "门禁产品合规",
    "门禁密码产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "门禁风险等级",
    "门禁风险分析",
    "门禁整改建议",
    "视频记录完整性",
    "视频算法",
    "视频技术合规",
    "视频产品合规",
    "视频密码产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "视频风险等级",
    "视频风险分析",
    "视频整改建议",
)

DEVICE_HEADERS = (
    "测评对象",
    "现场访谈记录",
    "身份鉴别方式",
    "密码技术使用",
    "算法",
    "技术合规",
    "产品合规",
    "身份鉴别密码产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "身份鉴别风险等级",
    "身份鉴别风险分析",
    "身份鉴别缓释机制",
    "身份鉴别缓释机制说明",
    "身份鉴别缓释后风险等级",
    "身份鉴别整改建议",
    "远程访问位置",
    "集中管控",
    "远程密码使用",
    "远程技术合规",
    "远程产品合规",
    "远程产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "远程风险等级",
    "远程风险分析",
    "远程缓释机制",
    "远程缓释机制说明",
    "远程缓释后风险等级",
    "远程整改建议",
    "通信协议",
    "证书算法",
    "证书来源",
    "起始日期",
    "失效日期",
    "机密性算法",
    "完整性算法",
    "其他",
    "访问控制完整性判定",
    "访问控制算法",
    "访问控制技术合规",
    "访问控制产品合规",
    "访问控制产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "访问控制风险等级",
    "访问控制风险分析",
    "访问控制整改建议",
    "日志完整性判定",
    "日志算法",
    "日志技术合规",
    "日志产品合规",
    "日志产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "日志风险等级",
    "日志风险分析",
    "日志整改建议",
    "程序完整性判定",
    "程序算法",
    "程序技术合规",
    "程序产品合规",
    "程序产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "程序风险等级",
    "程序风险分析",
    "程序整改建议",
)

NETWORK_HEADERS = (
    "子系统名称",
    "通信信道名称",
    "现场访谈记录",
    "所属子系统(选项)",
    "网络环境",
    "客户端形态",
    "服务端形态",
    "身份鉴别判定",
    "实现算法",
    "密码技术合规",
    "密码产品合规",
    "通信协议",
    "证书算法",
    "证书来源",
    "证书起始日期",
    "证书失效日期",
    "证书其他信息",
    "身份鉴别产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "身份鉴别风险等级",
    "身份鉴别风险分析",
    "身份鉴别缓释机制",
    "身份鉴别缓释机制说明",
    "身份鉴别缓释后风险等级",
    "身份鉴别整改建议",
    "完整性判定",
    "完整性算法",
    "完整性技术合规",
    "完整性产品合规",
    "完整性产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "完整性风险等级",
    "完整性风险分析",
    "完整性整改建议",
    "机密性判定",
    "机密性算法",
    "机密性技术合规",
    "机密性产品合规",
    "机密性产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "机密性风险等级",
    "机密性风险分析",
    "机密性缓释机制",
    "机密性缓释机制说明",
    "机密性缓释后风险等级",
    "机密性整改建议",
    "边界访问控制技术合规",
    "边界访问控制产品合规",
    "边界访问控制产品等级",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "边界访问控制风险等级",
    "边界访问控制风险分析",
    "边界访问控制整改建议",
)

APPLICATION_USER_HEADERS = (
    "子系统",
    "用户名称",
    "现场访谈记录",
    "身份鉴别方式",
    "密码技术使用",
    "实现算法",
    "密码技术合规",
    "密码产品合规",
    "密码产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "密钥管理情况",
    "风险等级",
    "风险分析",
    "缓释机制",
    "缓释机制说明",
    "缓释后风险等级",
    "整改建议",
)
APPLICATION_ACCESS_HEADERS = (
    "子系统",
    "访问控制信息名称",
    "现场访谈记录",
    "完整性判定",
    "实现算法",
    "密码技术合规",
    "密码产品合规",
    "密码产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "密钥管理情况",
    "风险等级",
    "风险分析",
    "整改建议",
)
APPLICATION_IMPORTANT_DATA_HEADERS = tuple(
    ["子系统", "重要数据名称", "数据类型", "现场访谈记录"]
    + _data_unit_headers("传输机密性", include_channel=True, include_mitigation=True)
    + _data_unit_headers("存储机密性", include_channel=False, include_mitigation=True)
    + _data_unit_headers("传输完整性", include_channel=True, include_mitigation=False)
    + _data_unit_headers("存储完整性", include_channel=False, include_mitigation=True)
)
APPLICATION_BUSINESS_HEADERS = (
    "子系统",
    "行为名称",
    "现场访谈记录",
    "不可否认性判定",
    "实现算法",
    "实现机制说明",
    "图片",
    "密码技术合规",
    "密码产品合规",
    "密码产品",
    "D",
    "A",
    "K",
    "Ra",
    "Rk",
    "密钥管理情况",
    "风险等级",
    "风险分析",
    "缓释机制",
    "缓释机制说明",
    "缓释后风险等级",
    "整改建议",
)


class ExportWriter:
    def __init__(self, engine) -> None:
        self.engine = engine
        self.basic_info = BasicInfoService(engine)
        self.physical = PhysicalService(engine)
        self.device = DeviceService(engine)
        self.network = NetworkService(engine)
        self.application = ApplicationService(engine)

    def write_system_basic_info(self, workbook: Workbook, project_id: int) -> Worksheet:
        sheet = workbook.create_sheet(SHEET_NAMES.basic_info)
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 50
        result = self.basic_info.load_full_basic_info(project_id)
        payload = result.payload if result.success else {}
        basic = dict(payload.get("basic_info", {}))
        system = dict(payload.get("system_info", {}))
        crypto = dict(payload.get("crypto_application_info", {}))
        subsystems = payload.get("subsystems", [])
        sheet.append(["一、项目基本信息", ""])
        for label, value in (
            ("流转单编号", basic.get("flow_no", "")),
            ("信息系统名称", basic.get("system_name", "")),
            ("系统负责人", basic.get("contact_name", "")),
            ("联系方式", basic.get("contact_phone", "")),
            ("评估人员", basic.get("assessor_name", "")),
            ("联系方式", basic.get("assessor_phone", "")),
            ("访谈时间", basic.get("interview_time", "")),
        ):
            sheet.append([label, self._date_text(value)])
        sheet.append(["", ""])
        sheet.append(["二、系统基本信息", ""])
        for label, key, fallback in (
            ("系统业务功能简介", "business_description", ""),
            ("子系统信息", "subsystem_description", "、".join(map(str, subsystems))),
            ("系统上线时间", "online_date", ""),
            ("等保定级情况", "classified_status", ""),
            ("系统等保等级", "security_level", ""),
            ("S值", "s_value", ""),
            ("A值", "a_value", ""),
            ("是否与密评等级一致", "level_consistency", ""),
            ("等保测评情况", "assessment_status", ""),
            ("等保测评机构", "assessment_org", ""),
            ("等保测评时间", "assessment_date", ""),
            ("等保测评结论", "assessment_result", ""),
            ("符合率", "pass_rate", ""),
            ("上游系统", "upstream_systems", ""),
            ("下游系统", "downstream_systems", ""),
            ("其他系统", "other_systems", ""),
        ):
            sheet.append([label, self._date_text(system.get(key, fallback))])
        sheet.append(["", ""])
        sheet.append(["三、密码应用情况", ""])
        for label, key in (
            ("上次密评情况", "last_assessment_status"),
            ("密码测评机构", "last_assessment_org"),
            ("密码测评时间", "last_assessment_date"),
            ("密码测评结论", "last_assessment_result"),
            ("分数", "last_assessment_score"),
            ("密码应用方案", "has_scheme"),
            ("是否经过评审", "is_reviewed"),
            ("方案评审方式", "review_method"),
            ("方案评审机构", "review_org"),
            ("方案评审时间", "review_date"),
        ):
            sheet.append([label, self._date_text(crypto.get(key, ""))])
        self._style_basic_sheet(sheet)
        return sheet

    def write_physical(self, workbook: Workbook, project_id: int) -> Worksheet:
        sheet = workbook.create_sheet(SHEET_NAMES.physical)
        sheet.append(PHYSICAL_HEADERS)
        for obj in self.physical.list_objects(project_id):
            details = self.physical.load_details(obj.id)
            sheet.append(
                [
                    details.object.name,
                    details.object.interview_record,
                    details.object.location,
                    details.object.access_control_system,
                    details.object.video_system,
                    *self._physical_auth_values(project_id, details.auth),
                    *self._physical_integrity_values(
                        project_id,
                        details.access_integrity,
                    ),
                    *self._physical_integrity_values(project_id, details.video_integrity),
                ]
            )
        self._style_table_sheet(sheet)
        return sheet

    def write_device(self, workbook: Workbook, project_id: int) -> Worksheet:
        sheet = workbook.create_sheet(SHEET_NAMES.device)
        sheet.append(DEVICE_HEADERS)
        for obj in self.device.list_objects(project_id):
            details = self.device.load_details(obj.id)
            sheet.append(
                [
                    details.object.name,
                    details.object.interview_record,
                    *self._device_auth_values(project_id, details.auth),
                    *self._device_remote_values(project_id, details.remote_management),
                    *self._device_integrity_values(project_id, details.access_integrity),
                    *self._device_integrity_values(project_id, details.log_integrity),
                    *self._device_integrity_values(project_id, details.executable_integrity),
                ]
            )
        self._style_table_sheet(sheet)
        return sheet

    def write_network(self, workbook: Workbook, project_id: int) -> Worksheet:
        sheet = workbook.create_sheet(SHEET_NAMES.network)
        sheet.append(NETWORK_HEADERS)
        for subsystem in self.network.list_subsystems(project_id):
            for channel in self.network.list_channels(subsystem.id):
                details = self.network.load_details(channel.id)
                sheet.append(
                    [
                        subsystem.name,
                        details.channel.name,
                        details.channel.interview_record,
                        *self._network_auth_values(project_id, subsystem.name, details),
                        *self._network_integrity_values(project_id, details.integrity),
                        *self._network_confidentiality_values(
                            project_id,
                            details.confidentiality,
                        ),
                        *self._network_boundary_values(project_id, details.boundary),
                    ]
                )
        self._style_table_sheet(sheet)
        return sheet

    def write_application(self, workbook: Workbook, project_id: int) -> Worksheet:
        sheet = workbook.create_sheet(SHEET_NAMES.application)
        subsystem_by_id = {
            subsystem.id: subsystem.name
            for subsystem in self.application.list_subsystems(project_id)
        }
        self._write_application_section(
            sheet,
            APPLICATION_SECTIONS.user_auth,
            APPLICATION_USER_HEADERS,
            self._application_user_rows(project_id, subsystem_by_id),
        )
        self._write_application_section(
            sheet,
            APPLICATION_SECTIONS.access_integrity,
            APPLICATION_ACCESS_HEADERS,
            self._application_access_rows(project_id, subsystem_by_id),
        )
        self._write_application_section(
            sheet,
            APPLICATION_SECTIONS.important_data,
            APPLICATION_IMPORTANT_DATA_HEADERS,
            self._application_data_rows(project_id, subsystem_by_id),
        )
        self._write_application_section(
            sheet,
            APPLICATION_SECTIONS.non_repudiation,
            APPLICATION_BUSINESS_HEADERS,
            self._application_business_rows(project_id, subsystem_by_id),
        )
        set_column_widths(sheet, [18] * max(sheet.max_column, 1))
        return sheet

    def _physical_auth_values(self, project_id: int, unit) -> list[object]:
        extra = self._mapping(unit.extra_data)
        return [
            unit.auth_methods,
            unit.crypto_usage,
            unit.algorithm,
            unit.compliance_status,
            unit.product_compliance,
            self._products_text(unit.products),
            *self._quant_values(project_id, unit.unit_type, unit.id),
            self._text(extra.get("guarded")),
            self._text(extra.get("registered")),
            self._text(extra.get("accompanied")),
            self._text(extra.get("realtime_monitoring")),
            unit.risk_level,
            unit.risk_analysis,
            self._text(extra.get("mitigation_available")),
            self._text(extra.get("mitigation_note")),
            self._text(extra.get("mitigated_level")),
            unit.remediation,
        ]

    def _physical_integrity_values(self, project_id: int, unit) -> list[object]:
        return [
            unit.implementation,
            unit.algorithm,
            unit.compliance_status,
            unit.product_compliance,
            self._products_text(unit.products),
            *self._quant_values(project_id, unit.unit_type, unit.id),
            unit.risk_level,
            unit.risk_analysis,
            unit.remediation,
        ]

    def _device_auth_values(self, project_id: int, unit) -> list[object]:
        extra = self._mapping(unit.extra_data)
        return [
            unit.auth_methods,
            unit.crypto_usage,
            unit.algorithm,
            unit.compliance_status,
            unit.product_compliance,
            self._products_text(unit.products),
            *self._quant_values(project_id, unit.unit_type, unit.id),
            unit.risk_level,
            unit.risk_analysis,
            self._text(extra.get("mitigation_available")),
            self._text(extra.get("mitigation_note")),
            self._text(extra.get("mitigated_level")),
            unit.remediation,
        ]

    def _device_remote_values(self, project_id: int, unit) -> list[object]:
        extra = self._mapping(unit.extra_data)
        return [
            self._text(extra.get("remote_position")),
            self._text(extra.get("centralized_management")),
            unit.crypto_usage,
            unit.compliance_status,
            unit.product_compliance,
            self._products_text(unit.products),
            *self._quant_values(project_id, unit.unit_type, unit.id),
            unit.risk_level,
            unit.risk_analysis,
            self._text(extra.get("mitigation_available")),
            self._text(extra.get("mitigation_note")),
            self._text(extra.get("mitigated_level")),
            unit.remediation,
            unit.remote_protocol,
            self._text(extra.get("certificate_algorithm")),
            self._text(extra.get("certificate_source")),
            self._date_text(extra.get("certificate_start_date")),
            self._date_text(extra.get("certificate_end_date")),
            self._text(extra.get("confidentiality_algorithm")),
            self._text(extra.get("integrity_algorithm")),
            self._text(extra.get("other_info")),
        ]

    def _device_integrity_values(self, project_id: int, unit) -> list[object]:
        extra = self._mapping(unit.extra_data)
        return [
            self._text(extra.get("implementation_status", unit.implementation)),
            unit.algorithm,
            unit.compliance_status,
            unit.product_compliance,
            self._products_text(unit.products),
            *self._quant_values(project_id, unit.unit_type, unit.id),
            unit.risk_level,
            unit.risk_analysis,
            unit.remediation,
        ]

    def _network_auth_values(self, project_id: int, subsystem_name: str, details) -> list[object]:
        unit = details.auth
        channel = details.channel
        extra = self._mapping(unit.extra_data)
        return [
            subsystem_name,
            channel.network_environment,
            channel.client_type,
            channel.server_type,
            self._text(extra.get("implementation_status", unit.implementation)),
            unit.algorithm,
            unit.compliance_status,
            unit.product_compliance,
            channel.protocol,
            self._text(extra.get("certificate_algorithm")),
            self._text(extra.get("certificate_source")),
            self._date_text(extra.get("certificate_start_date")),
            self._date_text(extra.get("certificate_end_date")),
            self._text(extra.get("certificate_other_info")),
            self._products_text(unit.products),
            *self._quant_values(project_id, unit.unit_type, unit.id),
            unit.risk_level,
            unit.risk_analysis,
            self._text(extra.get("mitigation_available")),
            self._text(extra.get("mitigation_note")),
            self._text(extra.get("mitigated_level")),
            unit.remediation,
        ]

    def _network_integrity_values(self, project_id: int, unit) -> list[object]:
        extra = self._mapping(unit.extra_data)
        return [
            self._text(extra.get("implementation_status", unit.implementation)),
            unit.algorithm,
            unit.compliance_status,
            unit.product_compliance,
            self._products_text(unit.products),
            *self._quant_values(project_id, unit.unit_type, unit.id),
            unit.risk_level,
            unit.risk_analysis,
            unit.remediation,
        ]

    def _network_confidentiality_values(self, project_id: int, unit) -> list[object]:
        extra = self._mapping(unit.extra_data)
        return [
            self._text(extra.get("implementation_status", unit.implementation)),
            unit.algorithm,
            unit.compliance_status,
            unit.product_compliance,
            self._products_text(unit.products),
            *self._quant_values(project_id, unit.unit_type, unit.id),
            unit.risk_level,
            unit.risk_analysis,
            self._text(extra.get("mitigation_available")),
            self._text(extra.get("mitigation_note")),
            self._text(extra.get("mitigated_level")),
            unit.remediation,
        ]

    def _network_boundary_values(self, project_id: int, unit) -> list[object]:
        return [
            unit.compliance_status,
            unit.product_compliance,
            unit.boundary_product_level,
            *self._quant_values(project_id, unit.unit_type, unit.id),
            unit.risk_level,
            unit.risk_analysis,
            unit.remediation,
        ]

    def _application_user_rows(
        self,
        project_id: int,
        subsystem_by_id: dict[int, str],
    ) -> list[list[object]]:
        rows = []
        for subsystem_id, subsystem_name in subsystem_by_id.items():
            for user in self.application.list_users(subsystem_id):
                detail = self.application.load_user_details(user.id)
                extra = self._mapping(detail.auth.extra_data)
                rows.append(
                    [
                        subsystem_name,
                        detail.user.name,
                        self._text(detail.user.extra_data.get("interview_record"))
                        if detail.user.extra_data
                        else "",
                        detail.auth.auth_methods,
                        detail.auth.crypto_usage,
                        detail.auth.algorithm,
                        detail.auth.compliance_status,
                        detail.auth.product_compliance,
                        self._products_text(detail.auth.products),
                        *self._quant_values(project_id, detail.auth.unit_type, detail.auth.id),
                        self._text(extra.get("key_management")),
                        detail.auth.risk_level,
                        detail.auth.risk_analysis,
                        self._text(extra.get("mitigation_available")),
                        self._text(extra.get("mitigation_note")),
                        self._text(extra.get("mitigated_level")),
                        detail.auth.remediation,
                    ]
                )
        return rows

    def _application_access_rows(
        self,
        project_id: int,
        subsystem_by_id: dict[int, str],
    ) -> list[list[object]]:
        rows = []
        for subsystem_id, subsystem_name in subsystem_by_id.items():
            for obj in self.application.list_access_controls(subsystem_id):
                detail = self.application.load_access_control_details(obj.id)
                extra = self._mapping(detail.integrity.extra_data)
                implementation = extra.get(
                    "implementation_status",
                    detail.integrity.implementation,
                )
                rows.append(
                    [
                        subsystem_name,
                        detail.access_control.name,
                        self._text(detail.access_control.extra_data.get("interview_record"))
                        if detail.access_control.extra_data
                        else "",
                        self._text(implementation),
                        detail.integrity.algorithm,
                        detail.integrity.compliance_status,
                        detail.integrity.product_compliance,
                        self._products_text(detail.integrity.products),
                        *self._quant_values(
                            project_id,
                            detail.integrity.unit_type,
                            detail.integrity.id,
                        ),
                        self._text(extra.get("key_management")),
                        detail.integrity.risk_level,
                        detail.integrity.risk_analysis,
                        detail.integrity.remediation,
                    ]
                )
        return rows

    def _application_data_rows(
        self,
        project_id: int,
        subsystem_by_id: dict[int, str],
    ) -> list[list[object]]:
        rows = []
        for subsystem_id, subsystem_name in subsystem_by_id.items():
            for obj in self.application.list_important_data(subsystem_id):
                details = self.application.load_important_data_details(obj.id)
                rows.append(
                    [
                        subsystem_name,
                        details.data.name,
                        details.data.data_type,
                        self._text(details.data.extra_data.get("interview_record"))
                        if details.data.extra_data
                        else "",
                        *self._application_data_unit_values(
                            project_id,
                            details.transport_confidentiality,
                            include_channel=True,
                            include_mitigation=True,
                        ),
                        *self._application_data_unit_values(
                            project_id,
                            details.storage_confidentiality,
                            include_channel=False,
                            include_mitigation=True,
                        ),
                        *self._application_data_unit_values(
                            project_id,
                            details.transport_integrity,
                            include_channel=True,
                            include_mitigation=False,
                        ),
                        *self._application_data_unit_values(
                            project_id,
                            details.storage_integrity,
                            include_channel=False,
                            include_mitigation=True,
                        ),
                    ]
                )
        return rows

    def _application_data_unit_values(
        self,
        project_id: int,
        unit,
        *,
        include_channel: bool,
        include_mitigation: bool,
    ) -> list[object]:
        extra = self._mapping(unit.extra_data)
        values: list[object] = [
            self._text(extra.get("implementation_status", unit.implementation)),
            unit.algorithm,
            self._text(extra.get("mechanism_description")),
            self._text(extra.get("image_path")),
            unit.compliance_status,
            unit.product_compliance,
            self._products_text(unit.products),
            *self._quant_values(project_id, unit.unit_type, unit.id),
        ]
        if include_channel:
            values.append(self._text(extra.get("related_channel")))
        values.extend(
            [
                self._text(extra.get("key_management")),
                unit.risk_level,
                unit.risk_analysis,
            ]
        )
        if include_mitigation:
            values.extend(
                [
                    self._text(extra.get("mitigation_available")),
                    self._text(extra.get("mitigation_note")),
                    self._text(extra.get("mitigated_level")),
                ]
            )
        values.append(unit.remediation)
        return values

    def _application_business_rows(
        self,
        project_id: int,
        subsystem_by_id: dict[int, str],
    ) -> list[list[object]]:
        rows = []
        for subsystem_id, subsystem_name in subsystem_by_id.items():
            for obj in self.application.list_business_actions(subsystem_id):
                detail = self.application.load_business_action_details(obj.id)
                extra = self._mapping(detail.non_repudiation.extra_data)
                implementation = extra.get(
                    "implementation_status",
                    detail.non_repudiation.implementation,
                )
                rows.append(
                    [
                        subsystem_name,
                        detail.action.name,
                        self._text(detail.action.extra_data.get("interview_record"))
                        if detail.action.extra_data
                        else "",
                        self._text(implementation),
                        detail.non_repudiation.algorithm,
                        self._text(extra.get("mechanism_description")),
                        self._text(extra.get("image_path")),
                        detail.non_repudiation.compliance_status,
                        detail.non_repudiation.product_compliance,
                        self._products_text(detail.non_repudiation.products),
                        *self._quant_values(
                            project_id,
                            detail.non_repudiation.unit_type,
                            detail.non_repudiation.id,
                        ),
                        self._text(extra.get("key_management")),
                        detail.non_repudiation.risk_level,
                        detail.non_repudiation.risk_analysis,
                        self._text(extra.get("mitigation_available")),
                        self._text(extra.get("mitigation_note")),
                        self._text(extra.get("mitigated_level")),
                        detail.non_repudiation.remediation,
                    ]
                )
        return rows

    def _write_application_section(
        self,
        sheet: Worksheet,
        title: str,
        headers: tuple[str, ...],
        rows: list[list[object]],
    ) -> None:
        if sheet.max_row > 1 or sheet.cell(row=1, column=1).value is not None:
            sheet.append([])
        title_row = sheet.max_row + 1
        sheet.append([title])
        header_row = sheet.max_row + 1
        sheet.append(headers)
        apply_header_row(sheet, title_row, max_column=max(1, len(headers)))
        apply_header_row(sheet, header_row, max_column=len(headers))
        for row in rows:
            sheet.append(row)
        apply_table_style(sheet, min_row=header_row + 1, max_col=len(headers))

    def _style_basic_sheet(self, sheet: Worksheet) -> None:
        apply_table_style(sheet)
        for row_index in (1, 10, 28):
            for cell in sheet[row_index]:
                font = copy(cell.font)
                font.bold = True
                cell.font = font

    def _style_table_sheet(self, sheet: Worksheet) -> None:
        apply_table_style(sheet)
        apply_header_row(sheet, 1)
        set_column_widths(sheet, [18] * max(sheet.max_column, 1))

    def _quant_values(self, project_id: int, unit_type: str, related_id: int) -> list[object]:
        quant = QuantService(self.engine, project_id=project_id).load(unit_type, related_id)
        if quant.record_id is None:
            return ["", "", "", "", ""]
        return [quant.d, quant.a, quant.k, quant.ra, quant.rk]

    def _products_text(self, products: Any) -> str:
        if not products:
            return ""
        parts = []
        for product in products:
            values = self._mapping(product)
            name = self._text(values.get("name", values.get("product_name", "")))
            if not name:
                continue
            parts.append(
                f"{name}("
                f"{self._text(values.get('vendor'))}, "
                f"证书:{self._text(values.get('certificate_no'))}, "
                f"等级:{self._text(values.get('level', values.get('product_level')))}, "
                f"用途:{self._text(values.get('usage'))})"
            )
        return "；".join(parts)

    def _date_text(self, value: Any) -> str:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return self._text(value)

    def _mapping(self, value: Any) -> dict[str, Any]:
        return dict(value) if isinstance(value, dict) else {}

    def _text(self, value: Any = "") -> str:
        return "" if value is None else str(value).strip()


def normalize_module_names(modules: list[str] | tuple[str, ...]) -> list[str]:
    normalized: list[str] = []
    for module in modules:
        sheet_name = MODULE_SHEETS.get(module, module)
        if sheet_name not in ALL_DATA_SHEETS:
            raise ValueError(f"未知导出模块: {module}")
        if sheet_name not in normalized:
            normalized.append(sheet_name)
    return normalized


assert len(PHYSICAL_HEADERS) == len(
    PHYSICAL_OBJECT_COLUMNS + PHYSICAL_AUTH_COLUMNS + PHYSICAL_INTEGRITY_COLUMNS * 2
)
assert len(DEVICE_HEADERS) == len(
    DEVICE_OBJECT_COLUMNS
    + DEVICE_AUTH_COLUMNS
    + DEVICE_REMOTE_COLUMNS
    + DEVICE_INTEGRITY_COLUMNS * 3
)
assert len(NETWORK_HEADERS) == len(
    NETWORK_OBJECT_COLUMNS
    + NETWORK_AUTH_COLUMNS
    + NETWORK_INTEGRITY_COLUMNS
    + NETWORK_CONFIDENTIALITY_COLUMNS
    + NETWORK_BOUNDARY_COLUMNS
)
assert len(APPLICATION_USER_HEADERS) == len(APPLICATION_USER_COLUMNS)
assert len(APPLICATION_ACCESS_HEADERS) == len(APPLICATION_ACCESS_COLUMNS)
assert len(APPLICATION_IMPORTANT_DATA_HEADERS) == len(APPLICATION_IMPORTANT_DATA_COLUMNS)
assert len(APPLICATION_BUSINESS_HEADERS) == len(APPLICATION_BUSINESS_COLUMNS)
