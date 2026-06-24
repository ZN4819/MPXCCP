from __future__ import annotations

from collections.abc import Iterable

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

TITLE_FILL = "4472C4"
HEADER_FILL = "D9E2F3"
NOT_APPLICABLE_FILL = "F2F2F2"
DEFAULT_FONT_NAME = "微软雅黑"

THIN_SIDE = Side(style="thin", color="808080")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
DEFAULT_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_table_style(
    sheet: Worksheet,
    *,
    min_row: int = 1,
    max_row: int | None = None,
    min_col: int = 1,
    max_col: int | None = None,
) -> None:
    max_row = sheet.max_row if max_row is None else max_row
    max_col = sheet.max_column if max_col is None else max_col
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.font = Font(name=DEFAULT_FONT_NAME, size=10)
            cell.alignment = DEFAULT_ALIGNMENT
            cell.border = THIN_BORDER


def apply_header_row(sheet: Worksheet, row_index: int, *, max_column: int | None = None) -> None:
    max_column = sheet.max_column if max_column is None else max_column
    fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    for cell in sheet[row_index][:max_column]:
        cell.font = Font(name=DEFAULT_FONT_NAME, size=10, bold=True)
        cell.fill = fill
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_title_range(sheet: Worksheet, cell_range: str) -> None:
    fill = PatternFill(fill_type="solid", fgColor=TITLE_FILL)
    for row in sheet[cell_range]:
        for cell in row:
            cell.font = Font(name=DEFAULT_FONT_NAME, size=12, bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER


def apply_not_applicable_row(sheet: Worksheet, row_index: int, *, max_column: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor=NOT_APPLICABLE_FILL)
    for cell in sheet[row_index][:max_column]:
        cell.fill = fill


def set_column_widths(sheet: Worksheet, widths: Iterable[float], *, start_column: int = 1) -> None:
    for offset, width in enumerate(widths, start=start_column):
        letter = get_column_letter(offset)
        sheet.column_dimensions[letter].width = width
