from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mpxccp.domain.constants import CHECK, CROSS, EMPTY, SLASH
from mpxccp.integration.excel.schema import MAX_WORKBOOK_SIZE_BYTES


@dataclass(frozen=True)
class ImportCellContext:
    module: str
    sheet: str
    row: int
    column: int | None = None
    field_name: str = ""

    def with_field(self, field_name: str, column: int | None = None) -> ImportCellContext:
        return replace(
            self,
            field_name=field_name,
            column=self.column if column is None else column,
        )

    def describe(self) -> str:
        parts = [self.module, self.sheet, f"第{self.row}行"]
        if self.column is not None:
            parts.append(f"第{self.column}列")
        if self.field_name:
            parts.append(self.field_name)
        return " / ".join(parts)


class ImportWorkbookError(ValueError):
    pass


class ImportDataError(ValueError):
    def __init__(self, message: str, context: ImportCellContext | None = None) -> None:
        self.context = context
        prefix = f"{context.describe()}: " if context is not None else ""
        super().__init__(f"{prefix}{message}")


class ImportReader:
    def load_workbook(self, source: str | Path | Workbook) -> Workbook:
        if isinstance(source, Workbook):
            return source
        path = Path(source)
        if path.suffix.lower() != ".xlsx":
            raise ImportWorkbookError("导入文件必须是 .xlsx 格式")
        if not path.exists():
            raise ImportWorkbookError(f"导入文件不存在: {path}")
        if path.stat().st_size > MAX_WORKBOOK_SIZE_BYTES:
            raise ImportWorkbookError("导入文件超过 50MB 限制")
        try:
            return load_workbook(path)
        except Exception as exc:  # pragma: no cover - openpyxl gives many exception types.
            raise ImportWorkbookError(f"工作簿打开失败: {exc}") from exc

    def cell(self, sheet: Worksheet, row: int, column: int) -> Any:
        try:
            return self.clean_cell(sheet.cell(row=row, column=column).value)
        except Exception:
            return EMPTY

    def row_values(self, sheet: Worksheet, row: int, columns: tuple[str, ...]) -> dict[str, Any]:
        return {
            name: self.cell(sheet, row, column_index)
            for column_index, name in enumerate(columns, start=1)
        }

    def clean_cell(self, value: Any) -> Any:
        if value is None:
            return EMPTY
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        lowered = text.lower()
        if ("enum" in lowered or "strenum" in lowered) and "." in text:
            return EMPTY
        return text

    def parse_date(self, value: Any) -> date | None:
        value = self.clean_cell(value)
        if isinstance(value, date):
            return value
        if not value:
            return None
        try:
            return datetime.strptime(str(value), "%Y-%m-%d").date()
        except ValueError:
            return None

    def parse_product_text(self, value: Any) -> list[dict[str, str]]:
        text = str(self.clean_cell(value) or "").strip()
        if not text:
            return []
        products: list[dict[str, str]] = []
        for raw_part in text.split("；"):
            part = raw_part.strip()
            if not part:
                continue
            products.append(self._parse_one_product(part))
        return products

    def parse_quant_values(
        self,
        *,
        d: Any,
        a: Any,
        k: Any,
        ra: Any,
        rk: Any,
        context: ImportCellContext,
    ) -> dict[str, str | float]:
        return {
            "d": self._parse_symbol(d, allow_slash=False),
            "a": self._parse_symbol(a, allow_slash=True),
            "k": self._parse_symbol(k, allow_slash=True),
            "ra": self._parse_float(ra, self._quant_context(context, "Ra", 3)),
            "rk": self._parse_float(rk, self._quant_context(context, "Rk", 4)),
        }

    def _parse_one_product(self, text: str) -> dict[str, str]:
        default = {
            "name": text,
            "vendor": "",
            "certificate_no": "",
            "level": "一级",
            "usage": "",
        }
        left_index = text.rfind("(")
        if left_index <= 0 or not text.endswith(")"):
            return default
        name = text[:left_index].strip()
        detail_text = text[left_index + 1 : -1].strip()
        if not name or not detail_text:
            return default
        product = {**default, "name": name}
        parts = [part.strip() for part in detail_text.split(",")]
        if parts:
            product["vendor"] = parts[0]
        for part in parts[1:]:
            if part.startswith("证书:"):
                product["certificate_no"] = part.removeprefix("证书:").strip()
            elif part.startswith("等级:"):
                level = part.removeprefix("等级:").strip()
                product["level"] = "一级" if self._looks_like_legacy_enum(level) else level
            elif part.startswith("用途:"):
                product["usage"] = part.removeprefix("用途:").strip()
        return product

    def _looks_like_legacy_enum(self, value: str) -> bool:
        lowered = value.lower()
        return ("enum" in lowered or "productlevel" in lowered) and "." in value

    def _parse_symbol(self, value: Any, *, allow_slash: bool) -> str:
        text = str(self.clean_cell(value) or "").strip()
        if text in {EMPTY, CHECK, CROSS}:
            return text
        if allow_slash and text == SLASH:
            return SLASH
        if text in {"√", "✓", "✔", "勾", "是", "通过", "符合", "1", "true", "True"}:
            return CHECK
        if text in {"×", "x", "X", "叉", "否", "不通过", "不符合", "0", "false", "False"}:
            return CROSS
        return EMPTY

    def _parse_float(self, value: Any, context: ImportCellContext) -> float:
        cleaned = self.clean_cell(value)
        if cleaned in (None, EMPTY, SLASH):
            return 1.0
        try:
            return float(cleaned)
        except (TypeError, ValueError) as exc:
            raise ImportDataError(f"{context.field_name} 必须是数字", context) from exc

    def _quant_context(
        self,
        context: ImportCellContext,
        field_name: str,
        offset: int,
    ) -> ImportCellContext:
        column = None if context.column is None else context.column + offset
        return context.with_field(field_name, column)
