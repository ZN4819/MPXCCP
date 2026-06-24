from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from sqlalchemy import delete

from mpxccp.domain.constants import CHECK
from mpxccp.domain.enums import SecurityLayer
from mpxccp.integration.excel.import_reader import ImportReader, ImportWorkbookError
from mpxccp.integration.excel.workbook_styles import (
    apply_header_row,
    apply_not_applicable_row,
    apply_table_style,
    apply_title_range,
    set_column_widths,
)
from mpxccp.models.scoring import ManagementScore
from mpxccp.repositories.scoring_repo import ScoringRepository
from mpxccp.repositories.session import session_scope
from mpxccp.services.result import ServiceResult
from mpxccp.services.scoring_service import (
    COMPLIANCE_SCORE,
    ScoreDetailRecord,
    ScoreSummaryRecord,
    ScoringService,
)

SCORE_SHEET_NAMES = [
    "整体测评",
    "1物理和环境安全",
    "2网络和通信安全",
    "3设备和计算安全",
    "4应用和数据安全",
    "5管理制度",
    "6人员管理",
    "7建设运行",
    "8应急处置",
]

TECHNICAL_SHEETS = {
    "1物理和环境安全": SecurityLayer.PHYSICAL.value,
    "2网络和通信安全": SecurityLayer.NETWORK.value,
    "3设备和计算安全": SecurityLayer.DEVICE.value,
    "4应用和数据安全": SecurityLayer.APPLICATION.value,
}
MANAGEMENT_SHEETS = {
    "5管理制度": SecurityLayer.MANAGEMENT_SYSTEM.value,
    "6人员管理": SecurityLayer.PERSONNEL.value,
    "7建设运行": SecurityLayer.CONSTRUCTION.value,
    "8应急处置": SecurityLayer.EMERGENCY.value,
}

OVERALL_HEADER_ROW_1 = [
    "序号",
    "层面（类）",
    "测评指标",
    "符合情况",
    "",
    "",
    "",
    "测评单元得分 Si,j",
    "指标权重 Wi,j",
    "安全层面得分情况 Si",
    "安全层面权重 Wi",
    "总分 S",
    "安全层面实际权重",
    "符合情况",
    "所占分值",
    "已得分值",
    "丢失分值",
]
OVERALL_HEADER_ROW_2 = ["", "", "", "符合", "部分符合", "不符合", "不适用"] + [""] * 10
TECHNICAL_HEADER_ROW_1 = [
    "测评单元",
    "测评对象",
    "适用情况",
    "涉及情况",
    "",
    "",
    "",
    "",
    "",
    "分值 Si,j,k",
    "符合情况",
    "测评单元得分 S i,j",
    "单元测评结果",
    "有效对象数量",
    "所占分值",
    "已得分值",
    "丢失分值",
]
TECHNICAL_HEADER_ROW_2 = [
    "",
    "",
    "",
    "密码使用有效性 D",
    "密码算法/技术合规性 A",
    "密钥管理安全 K",
    "Ra",
    "Rk",
    "得分",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
]

MANAGEMENT_RANGES = {
    SecurityLayer.MANAGEMENT_SYSTEM.value: range(23, 29),
    SecurityLayer.PERSONNEL.value: range(29, 34),
    SecurityLayer.CONSTRUCTION.value: range(34, 39),
    SecurityLayer.EMERGENCY.value: range(39, 42),
}


@dataclass(frozen=True)
class ScoreWorkbookImportResult:
    updated_count: int
    warnings: list[str]


class ScoreWorkbook:
    def __init__(
        self,
        engine,
        *,
        scoring_service: ScoringService | None = None,
        scoring_repo: ScoringRepository | None = None,
        reader: ImportReader | None = None,
    ) -> None:
        self.engine = engine
        self.scoring_service = scoring_service or ScoringService(engine)
        self.scoring_repo = scoring_repo or ScoringRepository()
        self.reader = reader or ImportReader()

    def export(self, project_id: int) -> Workbook:
        summary = self.scoring_service.refresh_all_technical_domains(project_id)
        workbook = Workbook()
        workbook.remove(workbook.active)
        self._write_overall(workbook, summary)
        for sheet_name, layer in TECHNICAL_SHEETS.items():
            self._write_technical_sheet(workbook, sheet_name, layer, summary)
        for sheet_name, layer in MANAGEMENT_SHEETS.items():
            self._write_management_sheet(workbook, sheet_name, layer, summary)
        return workbook

    def import_management_scores(
        self,
        project_id: int,
        source: str | Path | Workbook,
        *,
        mode: str,
    ) -> ServiceResult:
        if mode not in {"替换", "合并"}:
            return ServiceResult(
                success=False,
                message=f"不支持的打分表导入模式: {mode}",
                project_id=project_id,
            )
        try:
            workbook = self.reader.load_workbook(source)
        except ImportWorkbookError as exc:
            return ServiceResult(success=False, message=str(exc), project_id=project_id)

        try:
            result = self._import_management_workbook(project_id, workbook, mode=mode)
        except Exception as exc:
            return ServiceResult(
                success=False,
                message=f"打分表导入失败: {exc}",
                project_id=project_id,
            )
        return ServiceResult(
            success=True,
            message="打分表管理域评分导入完成",
            warnings=result.warnings,
            project_id=project_id,
            payload={"updated": result.updated_count},
        )

    def _write_overall(self, workbook: Workbook, summary: ScoreSummaryRecord) -> None:
        sheet = workbook.create_sheet(SCORE_SHEET_NAMES[0])
        sheet.append(OVERALL_HEADER_ROW_1)
        sheet.append(OVERALL_HEADER_ROW_2)
        for cell_range in (
            "A1:A2",
            "B1:B2",
            "C1:C2",
            "D1:G1",
            "H1:H2",
            "I1:I2",
            "J1:J2",
            "K1:K2",
            "L1:L2",
            "M1:M2",
            "N1:N2",
            "O1:O2",
            "P1:P2",
            "Q1:Q2",
        ):
            sheet.merge_cells(cell_range)
        detail_by_no = {detail.indicator_no: detail for detail in summary.details}
        layer_by_name = {layer.name: layer for layer in summary.layer_scores}
        for indicator_no in range(1, 42):
            detail = detail_by_no[indicator_no]
            layer = layer_by_name.get(detail.layer)
            row = [
                indicator_no,
                detail.layer,
                detail.indicator_name,
                CHECK if detail.compliance_status == "符合" else "",
                CHECK if detail.compliance_status == "部分符合" else "",
                CHECK if detail.compliance_status == "不符合" else "",
                CHECK if detail.compliance_status == "不适用" else "",
                self._score_value(detail.score),
                detail.indicator_weight,
                self._score_value(layer.score if layer else None),
                self._score_value(layer.actual_weight if layer else None),
                summary.total_score,
                self._score_value(layer.actual_weight if layer else None),
                detail.compliance_status,
                self._score_value(detail.allocated_score),
                self._score_value(detail.earned_score),
                self._score_value(detail.lost_score),
            ]
            sheet.append(row)
            if detail.not_applicable:
                apply_not_applicable_row(sheet, sheet.max_row, max_column=17)
        self._merge_overall_detail_ranges(sheet)
        total_row = sheet.max_row + 1
        sheet.append(
            [
                "合计",
                "",
                "",
                summary.compliant_count,
                summary.partial_count,
                summary.non_compliant_count,
                summary.not_applicable_count,
                41,
                "",
                "",
                "",
                summary.total_score,
                "",
                "",
                summary.total_allocated_score,
                summary.total_earned_score,
                summary.total_lost_score,
            ]
        )
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
        legend_row = sheet.max_row + 1
        sheet.append(
            [
                "符合情况",
                "",
                "",
                "符合",
                "部分符合",
                "不符合",
                "不适用",
                "综合得分",
                "",
                "",
                "",
                "",
            ]
        )
        sheet.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=3)
        sheet.merge_cells(start_row=legend_row, start_column=8, end_row=legend_row, end_column=12)
        apply_table_style(sheet)
        apply_header_row(sheet, 1, max_column=17)
        apply_header_row(sheet, 2, max_column=17)
        set_column_widths(sheet, [6, 16, 30, 6, 8, 6, 6, 10, 8, 10, 8, 8, 10, 8, 10, 10, 10])

    def _write_technical_sheet(
        self,
        workbook: Workbook,
        sheet_name: str,
        layer: str,
        summary: ScoreSummaryRecord,
    ) -> None:
        sheet = workbook.create_sheet(sheet_name)
        sheet.merge_cells("A1:L2")
        sheet["A1"] = layer
        apply_title_range(sheet, "A1:L2")
        sheet.append([])
        sheet.append(TECHNICAL_HEADER_ROW_1)
        sheet.append(TECHNICAL_HEADER_ROW_2)
        for cell_range in (
            "A3:A4",
            "B3:B4",
            "C3:C4",
            "D3:I3",
            "J3:J4",
            "K3:K4",
            "L3:L4",
            "M3:M4",
            "N3:N4",
            "O3:O4",
            "P3:P4",
        ):
            sheet.merge_cells(cell_range)
        for detail in [item for item in summary.details if item.layer == layer]:
            self._append_technical_detail(sheet, detail)
        apply_table_style(sheet)
        apply_header_row(sheet, 3, max_column=17)
        apply_header_row(sheet, 4, max_column=17)
        set_column_widths(sheet, [18, 18, 8, 8, 10, 8, 6, 6, 6, 8, 8, 10, 10, 8, 10, 10, 10])

    def _append_technical_detail(self, sheet, detail: ScoreDetailRecord) -> None:
        object_rows = detail.object_rows or []
        if detail.not_applicable or detail.score is None or not object_rows:
            sheet.append(
                [
                    detail.indicator_name,
                    "",
                    "不适用",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "N/A",
                    "不适用",
                    "N/A",
                    "不适用",
                    0,
                    self._score_value(detail.allocated_score),
                    self._score_value(detail.earned_score),
                    self._score_value(detail.lost_score),
                ]
            )
            apply_not_applicable_row(sheet, sheet.max_row, max_column=17)
            return
        start_row = sheet.max_row + 1
        for index, obj in enumerate(object_rows):
            sheet.append(
                [
                    detail.indicator_name if index == 0 else "",
                    obj.object_name,
                    "适用",
                    obj.d,
                    obj.a,
                    obj.k,
                    obj.ra,
                    obj.rk,
                    self._score_value(obj.score),
                    self._score_value(obj.score),
                    detail.compliance_status,
                    self._score_value(detail.score) if index == 0 else "",
                    detail.compliance_status if index == 0 else "",
                    detail.effective_object_count if index == 0 else "",
                    self._score_value(detail.allocated_score) if index == 0 else "",
                    self._score_value(detail.earned_score) if index == 0 else "",
                    self._score_value(detail.lost_score) if index == 0 else "",
                ]
            )
        end_row = sheet.max_row
        if end_row > start_row:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            sheet.merge_cells(start_row=start_row, start_column=12, end_row=end_row, end_column=12)
            sheet.merge_cells(start_row=start_row, start_column=13, end_row=end_row, end_column=13)

    def _write_management_sheet(
        self,
        workbook: Workbook,
        sheet_name: str,
        layer: str,
        summary: ScoreSummaryRecord,
    ) -> None:
        sheet = workbook.create_sheet(sheet_name)
        details = [item for item in summary.details if item.layer == layer]
        last_column = 2 + len(details)
        sheet.cell(row=1, column=1, value="序号")
        sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        sheet.cell(row=1, column=2, value="测评对象")
        sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=last_column)
        sheet.cell(row=1, column=3, value="测评指标符合情况（符合/部分符合/不符合/不适用）")
        for offset, detail in enumerate(details, start=3):
            sheet.cell(row=2, column=offset, value=detail.indicator_name)
            sheet.cell(row=3, column=offset, value=self._management_compliance_value(detail))
            sheet.cell(row=4, column=offset, value=self._score_value(detail.score))
            sheet.cell(row=5, column=offset, value=detail.compliance_status)
            sheet.cell(row=6, column=offset, value=self._score_value(detail.allocated_score))
            sheet.cell(row=7, column=offset, value=self._score_value(detail.earned_score))
            sheet.cell(row=8, column=offset, value=self._score_value(detail.lost_score))
        sheet.cell(row=3, column=1, value=1)
        sheet.cell(row=3, column=2, value="管理体系")
        for row, label in ((4, "测评单元得分"), (5, "单元测评结果")):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            sheet.cell(row=row, column=1, value=label)
        for row, label in ((6, "所占分值"), (7, "已得分值"), (8, "丢失分值")):
            sheet.cell(row=row, column=2, value=label)
        apply_table_style(sheet)
        apply_header_row(sheet, 1, max_column=last_column)
        apply_header_row(sheet, 2, max_column=last_column)
        set_column_widths(sheet, [6, 24] + [16] * len(details))

    def _import_management_workbook(
        self,
        project_id: int,
        workbook: Workbook,
        *,
        mode: str,
    ) -> ScoreWorkbookImportResult:
        warnings: list[str] = []
        updated_count = 0
        with session_scope(self.engine) as session:
            self.scoring_repo.ensure_indicators(session)
            indicators = {
                indicator.indicator_no: indicator
                for indicator in self.scoring_repo.list_indicators(session)
            }
            for sheet_name, layer in MANAGEMENT_SHEETS.items():
                if sheet_name not in workbook.sheetnames:
                    warnings.append(f"缺少工作表 {sheet_name}，已跳过")
                    continue
                sheet = workbook[sheet_name]
                if sheet.max_row < 3:
                    warnings.append(f"{sheet_name} 至少需要 3 行数据，已跳过")
                    continue
                indicator_nos = list(MANAGEMENT_RANGES[layer])
                if mode == "替换":
                    session.execute(
                        delete(ManagementScore).where(
                            ManagementScore.project_id == project_id,
                            ManagementScore.layer == layer,
                        )
                    )
                for offset, indicator_no in enumerate(indicator_nos, start=3):
                    status = self._parse_management_value(sheet.cell(row=3, column=offset).value)
                    if status is None:
                        continue
                    indicator = indicators[indicator_no]
                    self.scoring_repo.upsert_management_score(
                        session,
                        project_id=project_id,
                        indicator_no=indicator_no,
                        layer=indicator.layer,
                        compliance_status=status,
                        score=COMPLIANCE_SCORE[status],
                    )
                    updated_count += 1
            self.scoring_service.calculate_and_persist_summary_in_session(session, project_id)
        return ScoreWorkbookImportResult(updated_count=updated_count, warnings=warnings)

    def _merge_overall_detail_ranges(self, sheet) -> None:
        data_start = 3
        data_end = 43
        layer_start = data_start
        current_layer = sheet.cell(row=data_start, column=2).value
        for row_index in range(data_start + 1, data_end + 2):
            layer = sheet.cell(row=row_index, column=2).value if row_index <= data_end else None
            if layer == current_layer:
                continue
            if current_layer and row_index - layer_start > 1:
                sheet.merge_cells(
                    start_row=layer_start,
                    start_column=2,
                    end_row=row_index - 1,
                    end_column=2,
                )
            layer_start = row_index
            current_layer = layer
        sheet.merge_cells(start_row=data_start, start_column=12, end_row=data_end, end_column=12)

    def _parse_management_value(self, value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            numeric = float(value)
            if numeric == 1.0:
                return "符合"
            if numeric == 0.5:
                return "部分符合"
            if numeric == 0.0:
                return "不符合"
            return None
        text = str(value).strip()
        if not text:
            return None
        if text in COMPLIANCE_SCORE:
            return text
        if text in {"N/A", "NA", "n/a", "na", "/"}:
            return "不适用"
        if text in {"√", "1"}:
            return "符合"
        if text == "0.5":
            return "部分符合"
        if text in {"×", "x", "X", "0"}:
            return "不符合"
        return None

    def _score_value(self, value: float | None) -> float | str:
        return "N/A" if value is None else round(float(value), 4)

    def _management_compliance_value(self, detail: ScoreDetailRecord) -> str:
        return "N/A" if detail.score is None else detail.compliance_status
