from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from mpxccp.domain.issue_templates import (
    append_first_level_product_note,
    describe_product,
    effective_issue_risk,
    join_multi_select,
)
from mpxccp.services.application_service import ApplicationService
from mpxccp.services.basic_info_service import BasicInfoService
from mpxccp.services.device_service import DeviceService
from mpxccp.services.network_service import NetworkService
from mpxccp.services.physical_service import PhysicalService

ISSUE_SHEET_NAME = "问题清单"
ISSUE_HEADERS = (
    "测评层面",
    "测评要求",
    "问题编号",
    "系统名称",
    "被测对象名称",
    "现状描述",
    "问题说明",
    "风险分析与缓释机制",
    "风险等级",
    "整改建议",
    "说明",
    "主要责任部门（建议）",
)
ISSUE_COLUMN_WIDTHS = (15, 30, 10, 20, 20, 40, 40, 40, 10, 30, 10, 15)
RISK_FILL_COLORS = {
    "高风险": "FF0000",
    "中风险": "FFA500",
    "低风险": "FFFF00",
    "无风险": "00FF00",
    "不适用": "808080",
}
RISK_RANK = {"高风险": 5, "中风险": 4, "低风险": 3, "无风险": 2, "不适用": 1, "": 0}
HEADER_FILL = "D9E1F2"
DEFAULT_FONT_NAME = "微软雅黑"
THIN_SIDE = Side(style="thin", color="808080")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

REQUIREMENT_TEXT = {
    "应8.1 a）": "应采用密码技术对进入机房的人员进行身份鉴别，保证进出人员身份的真实性。",
    "应8.1 b）": "宜采用密码技术保证电子门禁系统进出记录数据的存储完整性。",
    "应8.1 c）": "宜采用密码技术保证视频监控音像记录数据的存储完整性。",
    "应8.2 a）": "应采用密码技术对通信实体进行身份鉴别，保证通信实体身份的真实性。",
    "应8.2 b) c）": "宜采用密码技术保证通信过程中重要数据的机密性和完整性。",
    "应8.2 d）": "宜采用密码技术保证网络边界访问控制信息的完整性。",
    "应8.3 a）": "应采用密码技术对登录设备的用户进行身份鉴别，保证用户身份的真实性。",
    "应8.3 b）": "远程管理设备时宜采用密码技术建立安全的信息传输通道。",
    "应8.3 c）": "宜采用密码技术保证设备访问控制信息的完整性。",
    "应8.3 d）": "宜采用密码技术保证设备日志记录的完整性。",
    "应8.3 e）": "宜采用密码技术保证重要可执行程序来源真实性和完整性。",
    "应8.4 a）": "应采用密码技术对登录应用的用户进行身份鉴别，保证应用用户身份的真实性。",
    "应8.4 b）": "宜采用密码技术保证应用系统访问控制信息的完整性。",
    "应8.4 c）": "宜采用密码技术保证重要数据传输过程中的机密性。",
    "应8.4 d）": "宜采用密码技术保证重要数据存储过程中的机密性。",
    "应8.4 e）": "宜采用密码技术保证重要数据传输过程中的完整性。",
    "应8.4 f）": "宜采用密码技术保证重要数据存储过程中的完整性。",
    "应8.4 h）": "宜采用密码技术保证关键业务行为的不可否认性。",
}


@dataclass(frozen=True)
class IssueRow:
    layer: str
    requirement: str
    system_name: str
    object_name: str
    current_state: str
    issue_description: str
    risk_analysis: str
    risk_level: str
    remediation: str
    note: str = ""
    department: str = ""
    has_source_content: bool = True


class IssueWorkbook:
    def __init__(self, engine) -> None:
        self.engine = engine
        self.basic_info = BasicInfoService(engine)
        self.physical = PhysicalService(engine)
        self.device = DeviceService(engine)
        self.network = NetworkService(engine)
        self.application = ApplicationService(engine)

    def export(self, project_id: int) -> Workbook:
        system_name = self._system_name(project_id)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = ISSUE_SHEET_NAME
        self._write_title_and_headers(sheet, system_name)
        for issue_no, row in enumerate(self._collect_rows(project_id, system_name), start=1):
            sheet.append(
                [
                    row.layer,
                    self._expand_requirement(row.requirement),
                    issue_no,
                    row.system_name or system_name,
                    row.object_name,
                    row.current_state,
                    row.issue_description,
                    row.risk_analysis,
                    row.risk_level,
                    row.remediation,
                    row.note,
                    row.department,
                ]
            )
            self._apply_data_row_style(sheet, sheet.max_row)
        self._merge_continuous_cells(sheet)
        return workbook

    def _collect_rows(self, project_id: int, system_name: str) -> list[IssueRow]:
        rows: list[IssueRow] = []
        rows.extend(self._physical_rows(project_id, system_name))
        rows.extend(self._network_rows(project_id, system_name))
        rows.extend(self._device_rows(project_id, system_name))
        rows.extend(self._application_rows(project_id, system_name))
        return rows

    def _physical_rows(self, project_id: int, system_name: str) -> list[IssueRow]:
        rows: list[IssueRow] = []
        for obj in self.physical.list_objects(project_id):
            details = self.physical.load_details(obj.id)
            rows.append(
                self._unit_row(
                    layer="物理和环境",
                    requirement="应8.1 a）",
                    system_name=system_name,
                    object_name=details.object.name,
                    unit=details.auth,
                    current_state=self._physical_auth_state(details.object.name, details.auth),
                    issue_description=self._physical_auth_description(
                        details.object.name,
                        details.auth,
                    ),
                )
            )
            rows.append(
                self._unit_row(
                    layer="物理和环境",
                    requirement="应8.1 b）",
                    system_name=system_name,
                    object_name=details.object.name,
                    unit=details.access_integrity,
                    current_state=self._physical_integrity_state(
                        details.object.name,
                        details.access_integrity,
                        "电子门禁记录数据的存储完整性",
                    ),
                    issue_description=self._physical_integrity_description(
                        details.object.name,
                        details.access_integrity,
                        "电子门禁进出记录数据",
                    ),
                )
            )
            rows.append(
                self._unit_row(
                    layer="物理和环境",
                    requirement="应8.1 c）",
                    system_name=system_name,
                    object_name=details.object.name,
                    unit=details.video_integrity,
                    current_state=self._physical_integrity_state(
                        details.object.name,
                        details.video_integrity,
                        "视频监控音像记录数据的存储完整性",
                    ),
                    issue_description=self._physical_integrity_description(
                        details.object.name,
                        details.video_integrity,
                        "视频监控音像记录数据",
                    ),
                )
            )
        return [row for row in rows if self._row_has_content(row)]

    def _network_rows(self, project_id: int, system_name: str) -> list[IssueRow]:
        rows: list[IssueRow] = []
        for subsystem in self.network.list_subsystems(project_id):
            for channel in self.network.list_channels(subsystem.id):
                details = self.network.load_details(channel.id)
                channel_name = details.channel.name
                network_environment = (
                    self._text(details.channel.network_environment) or "未记录网络环境"
                )
                protocol = self._text(details.channel.protocol) or "未记录"
                rows.append(
                    self._unit_row(
                        layer="网络和通信",
                        requirement="应8.2 a）",
                        system_name=subsystem.name,
                        object_name=channel_name,
                        unit=details.auth,
                        current_state=self._network_auth_state(
                            channel_name,
                            network_environment,
                            protocol,
                            details.auth,
                        ),
                        issue_description=self._network_auth_description(
                            channel_name,
                            protocol,
                            details.auth,
                        ),
                    )
                )
                rows.append(
                    self._combined_network_security_row(
                        subsystem.name,
                        channel_name,
                        details.channel.protocol,
                        details.integrity,
                        details.confidentiality,
                    )
                )
                rows.append(
                    self._unit_row(
                        layer="网络和通信",
                        requirement="应8.2 d）",
                        system_name=subsystem.name,
                        object_name=channel_name,
                        unit=details.boundary,
                        current_state=self._network_boundary_state(
                            channel_name,
                            details.boundary,
                        ),
                        issue_description=self._network_boundary_description(
                            channel_name,
                            details.channel.protocol,
                            details.boundary,
                        ),
                    )
                )
        return [row for row in rows if self._row_has_content(row)]

    def _device_rows(self, project_id: int, system_name: str) -> list[IssueRow]:
        rows: list[IssueRow] = []
        for obj in self.device.list_objects(project_id):
            details = self.device.load_details(obj.id)
            rows.append(
                self._unit_row(
                    layer="设备和计算",
                    requirement="应8.3 a）",
                    system_name=system_name,
                    object_name=details.object.name,
                    unit=details.auth,
                    current_state=self._device_auth_state(details.object.name, details.auth),
                    issue_description=self._device_auth_description(
                        details.object.name,
                        details.auth,
                    ),
                )
            )
            rows.append(
                self._unit_row(
                    layer="设备和计算",
                    requirement="应8.3 b）",
                    system_name=system_name,
                    object_name=details.object.name,
                    unit=details.remote_management,
                    current_state=self._device_remote_state(
                        details.object.name,
                        details.remote_management,
                    ),
                    issue_description=self._device_remote_description(
                        details.object.name,
                        details.remote_management,
                    ),
                )
            )
            for requirement, target, unit in (
                ("应8.3 c）", "设备访问控制信息", details.access_integrity),
                ("应8.3 d）", "设备日志记录", details.log_integrity),
                ("应8.3 e）", "重要可执行程序", details.executable_integrity),
            ):
                rows.append(
                    self._unit_row(
                        layer="设备和计算",
                        requirement=requirement,
                        system_name=system_name,
                        object_name=details.object.name,
                        unit=unit,
                        current_state=self._device_integrity_state(
                            details.object.name,
                            unit,
                            target,
                        ),
                        issue_description=self._device_integrity_description(
                            details.object.name,
                            unit,
                            target,
                        ),
                    )
                )
        return [row for row in rows if self._row_has_content(row)]

    def _application_rows(self, project_id: int, system_name: str) -> list[IssueRow]:
        rows: list[IssueRow] = []
        for subsystem in self.application.list_subsystems(project_id):
            for user in self.application.list_users(subsystem.id):
                details = self.application.load_user_details(user.id)
                rows.append(
                    self._unit_row(
                        layer="应用和数据",
                        requirement="应8.4 a）",
                        system_name=subsystem.name,
                        object_name=details.user.name,
                        unit=details.auth,
                        current_state=self._application_auth_state(
                            subsystem.name,
                            details.user.name,
                            details.auth,
                        ),
                        issue_description=self._application_auth_description(
                            details.user.name,
                            details.auth,
                        ),
                    )
                )
            for obj in self.application.list_access_controls(subsystem.id):
                details = self.application.load_access_control_details(obj.id)
                rows.append(
                    self._unit_row(
                        layer="应用和数据",
                        requirement="应8.4 b）",
                        system_name=subsystem.name,
                        object_name=details.access_control.name,
                        unit=details.integrity,
                        current_state=self._application_access_state(
                            subsystem.name,
                            details.access_control.name,
                            details.integrity,
                        ),
                        issue_description=self._application_access_description(
                            details.access_control.name,
                            details.integrity,
                        ),
                    )
                )
            for obj in self.application.list_important_data(subsystem.id):
                details = self.application.load_important_data_details(obj.id)
                for requirement, target, unit in (
                    ("应8.4 c）", "传输机密性", details.transport_confidentiality),
                    ("应8.4 d）", "存储机密性", details.storage_confidentiality),
                    ("应8.4 e）", "传输完整性", details.transport_integrity),
                    ("应8.4 f）", "存储完整性", details.storage_integrity),
                ):
                    rows.append(
                        self._unit_row(
                            layer="应用和数据",
                            requirement=requirement,
                            system_name=subsystem.name,
                            object_name=details.data.name,
                            unit=unit,
                            current_state=self._application_data_state(
                                subsystem.name,
                                details.data.name,
                                target,
                                unit,
                            ),
                            issue_description=self._application_data_description(
                                details.data.name,
                                target,
                                unit,
                            ),
                        )
                    )
            for obj in self.application.list_business_actions(subsystem.id):
                details = self.application.load_business_action_details(obj.id)
                rows.append(
                    self._unit_row(
                        layer="应用和数据",
                        requirement="应8.4 h）",
                        system_name=subsystem.name,
                        object_name=details.action.name,
                        unit=details.non_repudiation,
                        current_state=self._application_non_repudiation_state(
                            subsystem.name,
                            details.action.name,
                            details.non_repudiation,
                        ),
                        issue_description=self._application_non_repudiation_description(
                            details.action.name,
                            details.non_repudiation,
                        ),
                    )
                )
        return [row for row in rows if self._row_has_content(row)]

    def _combined_network_security_row(
        self,
        subsystem_name: str,
        channel_name: str,
        protocol: str,
        integrity,
        confidentiality,
    ) -> IssueRow:
        chosen = self._worst_unit([integrity, confidentiality])
        state_word = (
            "未实现"
            if any(self._is_unimplemented(unit) for unit in (integrity, confidentiality))
            else "已实现"
        )
        state = f"{channel_name}{state_word}通信过程中重要数据的机密性和完整性。"
        description = self._network_combined_description(
            channel_name,
            protocol,
            integrity,
            confidentiality,
        )
        return self._unit_row(
            layer="网络和通信",
            requirement="应8.2 b) c）",
            system_name=subsystem_name,
            object_name=channel_name,
            unit=chosen,
            current_state=state,
            issue_description=description,
            has_source_content=self._unit_has_content(integrity)
            or self._unit_has_content(confidentiality),
        )

    def _unit_row(
        self,
        *,
        layer: str,
        requirement: str,
        system_name: str,
        object_name: str,
        unit,
        current_state: str,
        issue_description: str,
        has_source_content: bool | None = None,
    ) -> IssueRow:
        extra = self._mapping(getattr(unit, "extra_data", None))
        risk_fields = {
            "risk_level": self._text(getattr(unit, "risk_level", "")),
            "mitigation_available": extra.get("mitigation_available"),
            "mitigated_level": self._text(extra.get("mitigated_level")),
        }
        return IssueRow(
            layer=layer,
            requirement=requirement,
            system_name=system_name,
            object_name=object_name,
            current_state=current_state,
            issue_description=issue_description,
            risk_analysis=self._risk_analysis(unit),
            risk_level=effective_issue_risk(risk_fields),
            remediation=self._text(getattr(unit, "remediation", "")),
            has_source_content=(
                self._unit_has_content(unit)
                if has_source_content is None
                else has_source_content
            ),
        )

    def _technology_description(
        self,
        subject: str,
        unit,
        *,
        success_text: str,
        failure_text: str,
    ) -> str:
        implementation = self._text(getattr(unit, "implementation", ""))
        crypto_usage = self._text(getattr(unit, "crypto_usage", ""))
        algorithm = self._text(getattr(unit, "algorithm", ""))
        products = getattr(unit, "products", None) or []
        product_description = describe_product(products)
        first_level_note = append_first_level_product_note(products)
        if implementation in {"未使用", "未实现", "不适用", "不涉及"} or crypto_usage in {
            "未使用",
            "未实现",
        }:
            return f"{subject}未使用密码技术，{failure_text}。"
        if algorithm:
            return (
                f"{subject}使用{algorithm}算法，{product_description}，"
                f"{success_text}。{first_level_note}"
            ).strip()
        return f"{subject}{product_description}，{success_text}。{first_level_note}".strip()

    def _risk_analysis(self, unit) -> str:
        extra = self._mapping(getattr(unit, "extra_data", None))
        risk_level = self._text(getattr(unit, "risk_level", ""))
        mitigation_available = extra.get("mitigation_available")
        mitigation_note = self._text(extra.get("mitigation_note"))
        if risk_level == "高风险" and self._has_mitigation(mitigation_available):
            return mitigation_note or self._text(getattr(unit, "risk_analysis", ""))
        return self._text(getattr(unit, "risk_analysis", ""))

    def _physical_auth_state(self, object_name: str, unit) -> str:
        extra = self._mapping(getattr(unit, "extra_data", None))
        auth_methods = join_multi_select(getattr(unit, "auth_methods", ""))
        return (
            f"{object_name}采用的身份鉴别方式为{auth_methods or '未记录'}。"
            f"{object_name}{self._yes_no(extra.get('realtime_monitoring'))}实时监控，"
            f"{self._yes_no(extra.get('guarded'))}专人值守，"
            f"进出时{self._yes_no(extra.get('registered'))}进行登记，"
            f"{self._yes_no(extra.get('accompanied'))}专人陪同。"
        )

    def _physical_auth_description(self, object_name: str, unit) -> str:
        if self._is_not_applicable(unit):
            return "该测评指标不适用。"
        if self._is_unimplemented(unit):
            return (
                f"{object_name}采取的身份鉴别方式未使用密码技术，"
                "无法有效保证进出人员身份的真实性。"
            )
        return (
            f"{object_name}采取的身份鉴别方式{self._technical_detail(unit)}，"
            f"{self._product_description(unit)}，采取{self._algorithm(unit)}算法完成进出人员的身份鉴别，"
            f"可以保证进出人员身份的真实性。{self._supplement_note(unit)}"
        ).strip()

    def _physical_integrity_state(self, object_name: str, unit, target: str) -> str:
        if self._is_not_applicable(unit):
            return f"{object_name}不适用{target}。"
        status = "未实现" if self._is_unimplemented(unit) else "已实现"
        return f"{object_name}{status}{target}。"

    def _physical_integrity_description(self, object_name: str, unit, target: str) -> str:
        if self._is_not_applicable(unit):
            return "该测评指标不适用。"
        if self._is_unimplemented(unit):
            return f"{object_name}对存储的{target}未使用存储完整性保护，存在数据被非法篡改的风险。"
        return (
            f"{object_name}对存储的{target}{self._technical_detail(unit)}，"
            f"{self._product_description(unit)}，使用{self._algorithm(unit)}进行存储完整性保护，"
            f"可以有效防止数据被非法篡改。{self._supplement_note(unit)}"
        ).strip()

    def _network_auth_state(
        self,
        channel_name: str,
        network_environment: str,
        protocol: str,
        unit,
    ) -> str:
        status = "未实现" if self._is_unimplemented(unit) else "已实现"
        if self._is_not_applicable(unit):
            status = "不适用"
        return (
            f"{channel_name}通过{network_environment}进行访问，"
            f"使用{protocol}协议，{status}通信实体身份鉴别。"
        )

    def _network_auth_description(self, channel_name: str, protocol: str, unit) -> str:
        if self._is_not_applicable(unit):
            return "该测评指标不适用。"
        if self._is_unimplemented(unit):
            return f"{channel_name}使用{protocol}协议，该协议无法保证通信实体身份的真实性。"
        extra = self._mapping(getattr(unit, "extra_data", None))
        return (
            f"{channel_name}使用{protocol}协议，该协议使用{self._algorithm(unit)}实现通信实体的身份鉴别，"
            f"{self._technical_detail(unit)}，{self._product_description(unit)}；"
            f"证书算法为{self._text(extra.get('certificate_algorithm')) or '无'}、"
            f"证书来源为{self._text(extra.get('certificate_source')) or '无'}，"
            f"证书失效日期为{self._text(extra.get('certificate_end_date')) or '无'}，"
            f"可以保证通信实体身份的真实性。{self._supplement_note(unit)}"
        ).strip()

    def _network_combined_description(
        self,
        channel_name: str,
        protocol: str,
        integrity,
        confidentiality,
    ) -> str:
        if self._is_unimplemented(integrity) or self._is_unimplemented(confidentiality):
            return (
                f"{channel_name}使用{protocol}协议，"
                "该协议无法保证通信过程中重要数据的机密性、完整性。"
            )
        products = list(getattr(confidentiality, "products", None) or []) + list(
            getattr(integrity, "products", None) or []
        )
        product_description = describe_product(products)
        supplement = append_first_level_product_note(products)
        return (
            f"{channel_name}使用{protocol}协议，{self._technical_detail(confidentiality)}"
            f"保证通信数据的完整性和机密性，{product_description}，"
            f"机密性采用{self._algorithm(confidentiality)}算法，"
            f"完整性采用{self._algorithm(integrity)}算法，可以保证通信过程中重要数据的机密性和完整性。"
            f"{supplement}"
        ).strip()

    def _network_boundary_state(self, channel_name: str, unit) -> str:
        if self._is_not_applicable(unit):
            return f"{channel_name}不适用网络边界访问控制信息的完整性。"
        status = "未实现" if self._is_unimplemented(unit) else "已实现"
        return f"{channel_name}{status}网络边界访问控制信息的完整性。"

    def _network_boundary_description(self, channel_name: str, protocol: str, unit) -> str:
        if self._is_not_applicable(unit):
            return "该测评指标不适用。"
        if self._is_unimplemented(unit):
            return (
                f"{channel_name}使用{protocol}协议，"
                "未使用密码产品保证网络边界访问控制信息的完整性。"
            )
        level = self._text(getattr(unit, "boundary_product_level", "")) or "无"
        return (
            f"{channel_name}使用{protocol}协议，{self._technical_detail(unit)}保证通信数据的完整性和机密性，"
            f"未使用合规的密码产品，密码产品等级为{level}，可以保证网络边界访问控制信息的完整性。"
        )

    def _device_auth_state(self, object_name: str, unit) -> str:
        methods = join_multi_select(getattr(unit, "auth_methods", ""))
        if self._is_not_applicable(unit):
            return f"{object_name}不适用用户身份鉴别指标。"
        usage = "未使用密码技术" if self._is_unimplemented(unit) else "已使用密码技术"
        return f"{object_name}使用了{methods or '未记录'}方式对登录设备用户进行身份鉴别，{usage}。"

    def _device_auth_description(self, object_name: str, unit) -> str:
        if self._is_not_applicable(unit):
            return "该测评指标不适用。"
        if self._is_unimplemented(unit):
            return (
                f"{object_name}采取的身份鉴别方式未使用密码技术，"
                "无法有效保证设备登录人员身份的真实性。"
            )
        return (
            f"{object_name}采取的身份鉴别方式{self._technical_detail(unit)}，"
            f"{self._product_description(unit)}，实现算法为{self._algorithm(unit)}算法，"
            f"可以保证登录设备用户身份鉴别的真实性。{self._supplement_note(unit)}"
        ).strip()

    def _device_remote_state(self, object_name: str, unit) -> str:
        remote_position = self._text(self._mapping(unit.extra_data).get("remote_position"))
        if remote_position == "本地" or remote_position.startswith("其他"):
            return f"{object_name}不适用远程管理通道。"
        usage = "未使用" if self._is_unimplemented(unit) else "已使用"
        return f"{object_name}远程访问，{usage}密码技术建立安全的远程管理通道。"

    def _device_remote_description(self, object_name: str, unit) -> str:
        extra = self._mapping(unit.extra_data)
        remote_position = self._text(extra.get("remote_position"))
        if remote_position == "本地" or remote_position.startswith("其他"):
            return f"{object_name}本地访问，该测评指标不适用。"
        if self._is_unimplemented(unit):
            return (
                f"{object_name}未使用密码技术建立安全的远程管理通道，"
                "无法保证远程管理设备时的信息传输安全。"
            )
        algorithm = self._text(extra.get("confidentiality_algorithm")) or self._text(
            extra.get("integrity_algorithm")
        )
        return (
            f"{object_name}建立的远程管理通道{self._technical_detail(unit)}，"
            f"{self._product_description(unit)}，"
            f"通信协议为{self._text(unit.remote_protocol) or '未记录'}，"
            f"实现算法为{algorithm or self._algorithm(unit)}算法，可以保证远程管理通道的安全。"
            f"{self._supplement_note(unit)}"
        ).strip()

    def _device_integrity_state(self, object_name: str, unit, target: str) -> str:
        if self._is_product_judgement(unit):
            return self._device_product_judgement_text(object_name, unit, target)
        if self._is_not_applicable(unit):
            return f"{object_name}不适用{target}完整性。"
        status = "未实现" if self._is_unimplemented(unit) else "已实现"
        return f"{object_name}{status}{target}完整性。"

    def _device_integrity_description(self, object_name: str, unit, target: str) -> str:
        if self._is_not_applicable(unit):
            return "该测评指标不适用。"
        if self._is_product_judgement(unit):
            return self._device_product_judgement_text(object_name, unit, target)
        if self._is_unimplemented(unit):
            return f"{object_name}未使用密码技术保证{target}完整性。"
        return (
            f"{object_name}{self._technical_detail(unit)}，"
            f"{self._product_description(unit)}{self._supplement_note(unit)}"
        ).strip()

    def _application_auth_state(self, subsystem_name: str, user_name: str, unit) -> str:
        methods = join_multi_select(getattr(unit, "auth_methods", ""))
        if self._is_not_applicable(unit):
            return "被测信息系统不涉及用户身份鉴别。"
        usage = "未使用密码技术" if self._is_unimplemented(unit) else "使用了密码技术"
        return (
            f"{subsystem_name}的{user_name}使用了{methods or '未记录'}"
            f"身份鉴别方式实现对登录用户的身份鉴别，{usage}。"
        )

    def _application_auth_description(self, user_name: str, unit) -> str:
        if self._is_not_applicable(unit):
            return "该测评指标不适用。"
        if self._is_unimplemented(unit):
            return (
                f"{user_name}采取的身份鉴别方式未使用密码技术，"
                "无法有效保证系统登录人员身份的真实性。"
            )
        return (
            f"{user_name}采取的身份鉴别方式{self._technical_detail(unit)}，"
            f"{self._product_description(unit)}，实现算法为{self._algorithm(unit)}算法，"
            f"可以保证登录设备用户身份鉴别的真实性。{self._supplement_note(unit)}"
        ).strip()

    def _application_access_state(self, subsystem_name: str, object_name: str, unit) -> str:
        if self._is_not_applicable(unit):
            return f"{subsystem_name}不涉及访问控制信息。"
        status = "未实现" if self._is_unimplemented(unit) else "已实现"
        return f"{subsystem_name}的{object_name}{status}存储完整性保护。"

    def _application_access_description(self, object_name: str, unit) -> str:
        if self._is_not_applicable(unit):
            return "该测评指标不适用。"
        if self._is_unimplemented(unit):
            return f"被测系统针对{object_name}未使用密码技术进行存储完整性保护。"
        return (
            f"被测系统针对{object_name}{self._technical_detail(unit)}，"
            f"{self._product_description(unit)}，实现算法为{self._algorithm(unit)}算法，"
            f"可以保证数据的存储完整性。{self._supplement_note(unit)}"
        ).strip()

    def _application_data_state(
        self,
        subsystem_name: str,
        data_name: str,
        protection_type: str,
        unit,
    ) -> str:
        if self._is_not_applicable(unit):
            return f"{subsystem_name}的{data_name}不适用{protection_type}。"
        status = "未实现" if self._is_unimplemented(unit) else "已实现"
        return f"{subsystem_name}的{data_name}{status}{protection_type}保护。"

    def _application_data_description(self, data_name: str, protection_type: str, unit) -> str:
        if self._is_not_applicable(unit):
            return "该测评指标不适用。"
        if self._is_unimplemented(unit):
            return f"被测系统针对{data_name}未使用密码技术进行{protection_type}保护。"
        return (
            f"被测系统针对{data_name}{self._technical_detail(unit)}，"
            f"{self._product_description(unit)}，实现算法为{self._algorithm(unit)}算法，"
            f"可以保证数据的{protection_type}。{self._supplement_note(unit)}"
        ).strip()

    def _application_non_repudiation_state(
        self,
        subsystem_name: str,
        action_name: str,
        unit,
    ) -> str:
        if self._is_not_applicable(unit):
            return f"{subsystem_name}不涉及不可否认性场景。"
        status = "未实现" if self._is_unimplemented(unit) else "已实现"
        return f"{subsystem_name}的{action_name}{status}不可否认性。"

    def _application_non_repudiation_description(self, action_name: str, unit) -> str:
        if self._is_not_applicable(unit):
            return "该测评指标不适用。"
        if self._is_unimplemented(unit):
            return (
                f"被测系统针对{action_name}未使用密码技术"
                "保证其数据原发行为与接收行为的不可否认性。"
            )
        return (
            f"被测系统针对{action_name}{self._technical_detail(unit)}，"
            f"{self._product_description(unit)}，实现算法为{self._algorithm(unit)}算法，"
            f"可以保证数据原发或接收行为的不可否认性。{self._supplement_note(unit)}"
        ).strip()

    def _worst_unit(self, units: list[Any]):
        return max(
            units,
            key=lambda unit: RISK_RANK.get(self._effective_risk_level(unit), 0),
        )

    def _effective_risk_level(self, unit) -> str:
        extra = self._mapping(getattr(unit, "extra_data", None))
        return effective_issue_risk(
            {
                "risk_level": self._text(getattr(unit, "risk_level", "")),
                "mitigation_available": extra.get("mitigation_available"),
                "mitigated_level": self._text(extra.get("mitigated_level")),
            }
        )

    def _is_unimplemented(self, unit) -> bool:
        extra = self._mapping(getattr(unit, "extra_data", None))
        implementation = self._text(
            extra.get(
                "implementation_status",
                getattr(unit, "implementation", "") or getattr(unit, "crypto_usage", ""),
            )
        )
        crypto_usage = self._text(getattr(unit, "crypto_usage", ""))
        return implementation in {"未使用", "未实现"} or crypto_usage in {"未使用", "未实现"}

    def _is_not_applicable(self, unit) -> bool:
        extra = self._mapping(getattr(unit, "extra_data", None))
        implementation = self._text(
            extra.get(
                "implementation_status",
                getattr(unit, "implementation", "") or getattr(unit, "crypto_usage", ""),
            )
        )
        return implementation in {"不适用", "不涉及"}

    def _technical_detail(self, unit) -> str:
        extra = self._mapping(getattr(unit, "extra_data", None))
        candidates = (
            extra.get("mechanism_description"),
            getattr(unit, "integrity_method", ""),
            getattr(unit, "encryption_method", ""),
            getattr(unit, "signature_method", ""),
            getattr(unit, "timestamp_method", ""),
            getattr(unit, "certificate_usage", ""),
            getattr(unit, "evaluation_result", ""),
            getattr(unit, "implementation", ""),
            getattr(unit, "crypto_usage", ""),
        )
        for value in candidates:
            text = self._text(value)
            if text:
                return text
        return "已使用密码技术"

    def _algorithm(self, unit) -> str:
        return self._text(getattr(unit, "algorithm", "")) or "未记录"

    def _product_description(self, unit) -> str:
        return describe_product(getattr(unit, "products", None) or [])

    def _supplement_note(self, unit) -> str:
        return append_first_level_product_note(getattr(unit, "products", None) or [])

    def _is_product_judgement(self, unit) -> bool:
        extra = self._mapping(getattr(unit, "extra_data", None))
        product_used = self._text(extra.get("product_used"))
        product_level = self._product_level(unit)
        return product_used in {"是", "密码产品"} or (not product_used and bool(product_level))

    def _product_level(self, unit) -> str:
        extra = self._mapping(getattr(unit, "extra_data", None))
        product_level = self._text(extra.get("product_level"))
        if product_level:
            return product_level
        return self._text(getattr(unit, "boundary_product_level", ""))

    def _device_product_judgement_text(self, object_name: str, unit, target: str) -> str:
        level = self._product_level(unit)
        if level == "一级":
            return (
                f"{object_name}是合规的密码产品，产品认证等级为一级，"
                "不符合三级系统使用二级及以上密码产品的要求。"
            )
        return (
            f"{object_name}是合规的密码产品，产品认证等级为{level or '无'}，"
            f"可以保证{target}的完整性。"
        )

    def _has_mitigation(self, value: object) -> bool:
        if isinstance(value, bool):
            return value
        return self._text(value) in {"具备", "是", "有", "true", "True", "1"}

    def _yes_no(self, value: object) -> str:
        if isinstance(value, bool):
            return "已" if value else "未"
        return "已" if self._text(value) in {"是", "有", "true", "True", "1"} else "未"

    def _implemented_text(self, unit) -> str:
        implementation = self._text(getattr(unit, "implementation", ""))
        extra = self._mapping(getattr(unit, "extra_data", None))
        implementation = self._text(extra.get("implementation_status", implementation))
        if implementation in {"已使用", "已实现"}:
            return "已实现"
        if implementation in {"未使用", "未实现"}:
            return "未实现"
        if implementation in {"不适用", "不涉及"}:
            return "不适用"
        return "已记录"

    def _unit_has_content(self, unit) -> bool:
        if unit is None:
            return False
        values = [
            getattr(unit, "implementation", ""),
            getattr(unit, "crypto_usage", ""),
            getattr(unit, "algorithm", ""),
            getattr(unit, "risk_level", ""),
            getattr(unit, "risk_analysis", ""),
            getattr(unit, "remediation", ""),
        ]
        extra = self._mapping(getattr(unit, "extra_data", None))
        return (
            any(self._text(value) for value in values)
            or bool(getattr(unit, "products", None))
            or bool(extra)
        )

    def _row_has_content(self, row: IssueRow) -> bool:
        return row.has_source_content and any(
            self._text(value)
            for value in (
                row.object_name,
                row.current_state,
                row.issue_description,
                row.risk_analysis,
                row.risk_level,
                row.remediation,
            )
        )

    def _write_title_and_headers(self, sheet: Worksheet, system_name: str) -> None:
        sheet.append([f"【三级标准】 {system_name} 密码应用安全性评估问题清单"])
        sheet.merge_cells("A1:L1")
        sheet.row_dimensions[1].height = 50
        title = sheet["A1"]
        title.font = Font(name=DEFAULT_FONT_NAME, size=14, bold=True)
        title.alignment = CENTER_WRAP
        title.border = THIN_BORDER
        sheet.append(list(ISSUE_HEADERS))
        for column, width in enumerate(ISSUE_COLUMN_WIDTHS, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width
        header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
        for cell in sheet[2]:
            cell.font = Font(name=DEFAULT_FONT_NAME, size=10, bold=True)
            cell.fill = header_fill
            cell.alignment = CENTER_WRAP
            cell.border = THIN_BORDER

    def _apply_data_row_style(self, sheet: Worksheet, row_index: int) -> None:
        for column in range(1, 13):
            cell = sheet.cell(row=row_index, column=column)
            cell.font = Font(name=DEFAULT_FONT_NAME, size=10)
            cell.border = THIN_BORDER
            cell.alignment = LEFT_WRAP if column in {6, 7, 8, 10} else CENTER_WRAP
        risk_cell = sheet.cell(row=row_index, column=9)
        color = RISK_FILL_COLORS.get(self._text(risk_cell.value))
        if color:
            risk_cell.fill = PatternFill(fill_type="solid", fgColor=color)

    def _merge_continuous_cells(self, sheet: Worksheet) -> None:
        if sheet.max_row < 4:
            return
        for column in (1, 2, 3, 4, 10):
            start = 3
            current_value = sheet.cell(row=start, column=column).value
            for row in range(4, sheet.max_row + 2):
                value = sheet.cell(row=row, column=column).value if row <= sheet.max_row else None
                if value == current_value:
                    continue
                if current_value not in (None, "") and row - start > 1:
                    sheet.merge_cells(
                        start_row=start,
                        start_column=column,
                        end_row=row - 1,
                        end_column=column,
                    )
                    merged = sheet.cell(row=start, column=column)
                    merged.alignment = LEFT_WRAP if column == 10 else CENTER_WRAP
                    for merged_row in range(start, row):
                        sheet.cell(row=merged_row, column=column).border = THIN_BORDER
                start = row
                current_value = value

    def _expand_requirement(self, value: str) -> str:
        return REQUIREMENT_TEXT.get(value, value)

    def _system_name(self, project_id: int) -> str:
        result = self.basic_info.load_basic_info(project_id)
        if not result.success:
            return ""
        return self._text(result.payload.get("system_name"))

    def _mapping(self, value: Any) -> dict[str, Any]:
        return dict(value) if isinstance(value, dict) else {}

    def _text(self, value: Any = "") -> str:
        return "" if value is None else str(value).strip()
