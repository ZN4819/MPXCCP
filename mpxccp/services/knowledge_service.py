from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from sqlalchemy import Engine

from mpxccp.domain.enums import KnowledgeModule
from mpxccp.integration.excel.import_reader import ImportReader
from mpxccp.integration.excel.schema import KNOWLEDGE_HEADERS, KNOWLEDGE_SHEET_NAME
from mpxccp.integration.excel.workbook_styles import (
    apply_header_row,
    apply_table_style,
    set_column_widths,
)
from mpxccp.models.knowledge import KnowledgeEntry
from mpxccp.repositories.knowledge_repo import KnowledgeRepository
from mpxccp.repositories.session import readonly_session_scope, session_scope
from mpxccp.services.result import ServiceResult

DEFAULT_VISIBLE_MODULES = {
    KnowledgeModule.COMMON.value,
    KnowledgeModule.OTHER.value,
}


@dataclass(frozen=True)
class KnowledgeRecord:
    id: int
    entry_type: str
    module: str
    title: str
    content: str
    risk_level: str
    tags: str
    sort_order: int
    created_at: str = ""
    updated_at: str = ""


class KnowledgeService:
    def __init__(
        self,
        engine: Engine,
        knowledge_repo: KnowledgeRepository | None = None,
    ) -> None:
        self.engine = engine
        self.knowledge_repo = knowledge_repo or KnowledgeRepository()

    def list_entries(
        self,
        type: str,
        module: str,
        show_all: bool = False,
        text_filter: str = "",
    ) -> list[KnowledgeRecord]:
        modules = None if show_all else {module, *DEFAULT_VISIBLE_MODULES}
        with readonly_session_scope(self.engine) as session:
            return [
                self._record_payload(item)
                for item in self.knowledge_repo.list_entries(
                    session,
                    entry_type=type,
                    modules=modules,
                    text_filter=text_filter.strip(),
                )
            ]

    def list_all_entries(self) -> list[KnowledgeRecord]:
        with readonly_session_scope(self.engine) as session:
            return [
                self._record_payload(item)
                for item in self.knowledge_repo.list_all_entries(session)
            ]

    def add_entry(self, type: str, module: str, content: str) -> ServiceResult:
        cleaned = self._text(content)
        if not cleaned:
            return ServiceResult(
                success=False,
                message="knowledge content is empty",
                warnings=["empty_content"],
            )
        with session_scope(self.engine) as session:
            entry = self.knowledge_repo.add_entry(
                session,
                entry_type=self._text(type),
                module=self._text(module),
                content=cleaned,
                sort_order=self._next_sort_order(session, type),
            )
            return ServiceResult(
                success=True,
                message="knowledge entry added",
                payload={"entry_id": entry.id},
            )

    def update_entry(self, id: int, content: str, module: str) -> ServiceResult:
        cleaned = self._text(content)
        if not cleaned:
            return ServiceResult(
                success=False,
                message="knowledge content is empty",
                warnings=["empty_content"],
            )
        with session_scope(self.engine) as session:
            entry = session.get(KnowledgeEntry, id)
            if entry is None:
                return ServiceResult(
                    success=False,
                    message="knowledge entry not found",
                    warnings=["entry_not_found"],
                )
            entry.content = cleaned
            entry.title = cleaned[:60]
            entry.module = self._text(module)
            return ServiceResult(success=True, message="knowledge entry updated")

    def delete_entries(self, ids: list[int]) -> ServiceResult:
        with session_scope(self.engine) as session:
            deleted = self.knowledge_repo.delete_entries(session, ids)
            return ServiceResult(
                success=True,
                message="knowledge entries deleted",
                payload={"deleted": deleted},
            )

    def dedupe_append(self, entries: list[dict[str, Any]]) -> ServiceResult:
        added = 0
        skipped = 0
        with session_scope(self.engine) as session:
            for values in entries:
                entry_type = self._text(values.get("type", values.get("entry_type", "")))
                module = self._text(values.get("module", ""))
                content = self._text(values.get("content", ""))
                if not entry_type or not module or not content:
                    skipped += 1
                    continue
                duplicate = self.knowledge_repo.find_duplicate(
                    session,
                    entry_type=entry_type,
                    module=module,
                    content=content,
                )
                if duplicate is not None:
                    skipped += 1
                    continue
                self.knowledge_repo.add_entry(
                    session,
                    entry_type=entry_type,
                    module=module,
                    content=content,
                    sort_order=self._next_sort_order(session, entry_type),
                )
                added += 1
            return ServiceResult(
                success=True,
                message="knowledge entries appended",
                payload={"added": added, "skipped": skipped},
            )

    def replace_all(self, entries: list[dict[str, Any]]) -> ServiceResult:
        cleaned_entries = self._clean_entries(entries)
        if not cleaned_entries:
            return ServiceResult(
                success=False,
                message="no valid knowledge entries to replace",
                warnings=["no_valid_entries"],
                payload={"deleted": 0, "added": 0},
            )
        with session_scope(self.engine) as session:
            deleted = self.knowledge_repo.delete_all(session)
            for sort_order, values in enumerate(cleaned_entries):
                self.knowledge_repo.add_entry(
                    session,
                    entry_type=values["entry_type"],
                    module=values["module"],
                    content=values["content"],
                    sort_order=sort_order,
                )
            return ServiceResult(
                success=True,
                message="knowledge entries replaced",
                payload={"deleted": deleted, "added": len(cleaned_entries)},
            )

    def export_workbook(self) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = KNOWLEDGE_SHEET_NAME
        sheet.append(list(KNOWLEDGE_HEADERS))
        for entry in self.list_all_entries():
            sheet.append(
                [
                    entry.id,
                    entry.entry_type,
                    entry.module,
                    entry.content,
                    entry.created_at,
                    entry.updated_at,
                ]
            )
        apply_table_style(sheet)
        apply_header_row(sheet, 1)
        set_column_widths(sheet, [10, 16, 16, 60, 20, 20])
        return workbook

    def import_workbook(
        self,
        source: str | Path | Workbook,
        mode: str,
    ) -> ServiceResult:
        try:
            workbook = ImportReader().load_workbook(source)
            if KNOWLEDGE_SHEET_NAME not in workbook:
                return ServiceResult(
                    success=False,
                    message=f"knowledge workbook missing sheet: {KNOWLEDGE_SHEET_NAME}",
                    warnings=["knowledge_sheet_not_found"],
                )
            sheet = workbook[KNOWLEDGE_SHEET_NAME]
            headers = [self._text(sheet.cell(row=1, column=col).value) for col in range(1, 7)]
            if headers != list(KNOWLEDGE_HEADERS):
                return ServiceResult(
                    success=False,
                    message="knowledge workbook headers do not match expected schema",
                    warnings=["knowledge_headers_mismatch"],
                )
            entries = self._workbook_entries(sheet)
            if mode == "替换":
                return self.replace_import_entries(entries)
            if mode == "追加":
                return self.dedupe_append(entries)
            return ServiceResult(
                success=False,
                message=f"unsupported knowledge import mode: {mode}",
                warnings=["unsupported_import_mode"],
            )
        except Exception as exc:
            return ServiceResult(
                success=False,
                message=f"knowledge workbook import failed: {exc}",
                warnings=["knowledge_import_failed"],
            )

    def replace_import_entries(self, entries: list[dict[str, Any]]) -> ServiceResult:
        cleaned_entries = self._clean_entries(entries)
        skipped = len(entries) - len(cleaned_entries)
        with session_scope(self.engine) as session:
            deleted = self.knowledge_repo.delete_all(session)
            for sort_order, values in enumerate(cleaned_entries):
                self.knowledge_repo.add_entry(
                    session,
                    entry_type=values["entry_type"],
                    module=values["module"],
                    content=values["content"],
                    sort_order=sort_order,
                )
            return ServiceResult(
                success=True,
                message="knowledge entries replaced",
                payload={
                    "deleted": deleted,
                    "added": len(cleaned_entries),
                    "skipped": skipped,
                },
            )

    def _workbook_entries(self, sheet) -> list[dict[str, str]]:
        reader = ImportReader()
        entries: list[dict[str, str]] = []
        for row_index in range(2, sheet.max_row + 1):
            entry_type = self._text(reader.cell(sheet, row_index, 2))
            module = self._text(reader.cell(sheet, row_index, 3))
            content = self._text(reader.cell(sheet, row_index, 4))
            entries.append(
                {
                    "entry_type": entry_type,
                    "module": module,
                    "content": content,
                }
            )
        return entries

    def _clean_entries(self, entries: list[dict[str, Any]]) -> list[dict[str, str]]:
        cleaned_entries: list[dict[str, str]] = []
        for values in entries:
            entry_type = self._text(values.get("type", values.get("entry_type", "")))
            module = self._text(values.get("module", ""))
            content = self._text(values.get("content", ""))
            if not entry_type or not module or not content:
                continue
            cleaned_entries.append(
                {
                    "entry_type": entry_type,
                    "module": module,
                    "content": content,
                }
            )
        return cleaned_entries

    def _next_sort_order(self, session, entry_type: str) -> int:
        entries = self.knowledge_repo.list_entries(
            session,
            entry_type=entry_type,
            modules=None,
        )
        return len(entries)

    def _text(self, value: Any) -> str:
        return "" if value is None else str(value).strip()

    def _time_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.isoformat()
        return self._text(value)

    def _record_payload(self, entry: KnowledgeEntry) -> KnowledgeRecord:
        return KnowledgeRecord(
            id=entry.id,
            entry_type=entry.entry_type,
            module=entry.module,
            title=entry.title,
            content=entry.content,
            risk_level=entry.risk_level,
            tags=entry.tags,
            sort_order=entry.sort_order,
            created_at=self._time_text(entry.created_at),
            updated_at=self._time_text(entry.updated_at),
        )
