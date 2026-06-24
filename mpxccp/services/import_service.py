from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl.workbook import Workbook
from sqlalchemy import func, select, update
from sqlalchemy.orm import Session

import mpxccp.models as models
from mpxccp.domain.quant_rules import calculate_object_score
from mpxccp.integration.excel.import_reader import (
    ImportCellContext,
    ImportDataError,
    ImportReader,
    ImportWorkbookError,
)
from mpxccp.integration.excel.schema import (
    APPLICATION_ACCESS_COLUMNS,
    APPLICATION_BUSINESS_COLUMNS,
    APPLICATION_SECTIONS,
    APPLICATION_USER_COLUMNS,
    DEVICE_AUTH_COLUMNS,
    DEVICE_INTEGRITY_COLUMNS,
    DEVICE_OBJECT_COLUMNS,
    DEVICE_REMOTE_COLUMNS,
    IMPORTANT_DATA_STORAGE_COLUMNS,
    IMPORTANT_DATA_TRANSPORT_COLUMNS,
    IMPORTANT_DATA_TRANSPORT_INTEGRITY_COLUMNS,
    NETWORK_AUTH_COLUMNS,
    NETWORK_BOUNDARY_COLUMNS,
    NETWORK_COMMON_UNIT_COLUMNS,
    NETWORK_OBJECT_COLUMNS,
    PHYSICAL_AUTH_COLUMNS,
    PHYSICAL_INTEGRITY_COLUMNS,
    PHYSICAL_OBJECT_COLUMNS,
    SHEET_NAMES,
)
from mpxccp.models.shared import QuantitativeAssessment
from mpxccp.repositories.application_repo import ApplicationRepository
from mpxccp.repositories.basic_info_repo import BasicInfoRepository
from mpxccp.repositories.device_repo import DeviceRepository
from mpxccp.repositories.network_repo import NetworkRepository
from mpxccp.repositories.physical_repo import PhysicalRepository
from mpxccp.repositories.session import session_scope
from mpxccp.repositories.shared_repo import SharedRepository
from mpxccp.services.application_service import UNIT_CONFIG as APP_UNITS
from mpxccp.services.device_service import UNIT_CONFIG as DEVICE_UNITS
from mpxccp.services.migration_service import MigrationService
from mpxccp.services.network_service import UNIT_CONFIG as NETWORK_UNITS
from mpxccp.services.physical_service import UNIT_CONFIG as PHYSICAL_UNITS
from mpxccp.services.result import ServiceResult


class ImportService:
    def __init__(
        self,
        engine,
        *,
        reader: ImportReader | None = None,
        basic_repo: BasicInfoRepository | None = None,
        physical_repo: PhysicalRepository | None = None,
        device_repo: DeviceRepository | None = None,
        network_repo: NetworkRepository | None = None,
        application_repo: ApplicationRepository | None = None,
        shared_repo: SharedRepository | None = None,
    ) -> None:
        self.engine = engine
        self.reader = reader or ImportReader()
        self.basic_repo = basic_repo or BasicInfoRepository()
        self.physical_repo = physical_repo or PhysicalRepository()
        self.device_repo = device_repo or DeviceRepository()
        self.network_repo = network_repo or NetworkRepository()
        self.application_repo = application_repo or ApplicationRepository()
        self.shared_repo = shared_repo or SharedRepository()

    def detect_import_mode(self, project_id: int, workbook: Workbook) -> str:
        _ = project_id
        replace_sheets = {
            SHEET_NAMES.basic_info,
            SHEET_NAMES.physical,
            SHEET_NAMES.device,
        }
        if any(name in workbook.sheetnames for name in replace_sheets):
            return "替换"
        if any(
            name in workbook.sheetnames for name in (SHEET_NAMES.network, SHEET_NAMES.application)
        ):
            return "追加"
        return "替换"

    def import_interview_template(
        self,
        project_id: int,
        source: str | Path | Workbook,
        mode: str | None = None,
    ) -> ServiceResult:
        try:
            workbook = self.reader.load_workbook(source)
        except ImportWorkbookError as exc:
            return ServiceResult(success=False, message=str(exc), project_id=project_id)

        import_mode = mode or self.detect_import_mode(project_id, workbook)
        if import_mode not in {"替换", "追加"}:
            return ServiceResult(
                success=False,
                message=f"不支持的导入模式: {import_mode}",
                project_id=project_id,
            )

        warnings: list[str] = []
        if import_mode == "追加":
            warnings.append("追加模式：仅导入网络和应用模块的新子系统数据")

        try:
            with session_scope(self.engine) as session:
                if self.basic_repo.get_project(session, project_id) is None:
                    return ServiceResult(
                        success=False,
                        message="project not found",
                        warnings=["project_not_found"],
                        project_id=project_id,
                    )
                existing_subsystems = set(self._enabled_basic_subsystem_names(session, project_id))
                if import_mode == "替换":
                    if SHEET_NAMES.basic_info in workbook.sheetnames:
                        self.import_system_basic_info(
                            session,
                            project_id,
                            workbook[SHEET_NAMES.basic_info],
                        )
                    else:
                        warnings.append("未找到'系统基本信息'工作表，已跳过")
                    if SHEET_NAMES.physical in workbook.sheetnames:
                        self.import_physical(
                            session,
                            project_id,
                            workbook[SHEET_NAMES.physical],
                            replace=True,
                        )
                    if SHEET_NAMES.device in workbook.sheetnames:
                        self.import_device(
                            session,
                            project_id,
                            workbook[SHEET_NAMES.device],
                            replace=True,
                        )
                if SHEET_NAMES.network in workbook.sheetnames:
                    self.import_network(
                        session,
                        project_id,
                        workbook[SHEET_NAMES.network],
                        replace=import_mode == "替换",
                        skip_subsystems=existing_subsystems if import_mode == "追加" else set(),
                    )
                if SHEET_NAMES.application in workbook.sheetnames:
                    self.import_application(
                        session,
                        project_id,
                        workbook[SHEET_NAMES.application],
                        replace=import_mode == "替换",
                        skip_subsystems=existing_subsystems if import_mode == "追加" else set(),
                    )
                MigrationService(self.engine).clean_enum_values(session)
        except ImportDataError as exc:
            return ServiceResult(success=False, message=str(exc), project_id=project_id)
        except Exception as exc:
            return ServiceResult(
                success=False,
                message=f"导入异常: {exc}",
                project_id=project_id,
            )

        return ServiceResult(
            success=True,
            message="访谈模板导入完成",
            warnings=warnings,
            project_id=project_id,
        )

    def import_system_basic_info(self, session: Session, project_id: int, sheet) -> None:
        basic_values = {
            "flow_no": self.reader.cell(sheet, 2, 2),
            "system_name": self.reader.cell(sheet, 3, 2),
            "contact_name": self.reader.cell(sheet, 4, 2),
            "contact_phone": self.reader.cell(sheet, 5, 2),
            "assessor_name": self.reader.cell(sheet, 6, 2),
            "assessor_phone": self.reader.cell(sheet, 7, 2),
            "interview_time": self._date_text(self.reader.cell(sheet, 8, 2)),
        }
        project = self.basic_repo.get_project(session, project_id)
        if project is None:
            return
        current_basic = self.basic_repo.get_basic_info(session, project_id)
        self.basic_repo.upsert_basic_info(
            session,
            project,
            flow_no=str(basic_values["flow_no"] or project.flow_no),
            system_name=str(basic_values["system_name"] or project.system_name),
            client_name=project.client_name,
            assessment_org=project.assessment_org,
            contact_name=str(basic_values["contact_name"] or ""),
            contact_phone=str(basic_values["contact_phone"] or ""),
            interview_time=str(
                basic_values["interview_time"]
                or (current_basic.assessment_date if current_basic is not None else "")
            ),
            notes=current_basic.notes if current_basic is not None else "",
        )
        project.extra_data = {
            **self._mapping(project.extra_data),
            "assessor_name": str(basic_values["assessor_name"] or ""),
            "assessor_phone": str(basic_values["assessor_phone"] or ""),
            "interview_time": str(basic_values["interview_time"] or ""),
        }
        section_row = self._find_row(sheet, "二、系统基本信息")
        if section_row is None:
            return
        system_values = {
            "business_description": self.reader.cell(sheet, section_row + 1, 2),
            "subsystem_description": self.reader.cell(sheet, section_row + 2, 2),
            "online_date": self._date_text(self.reader.cell(sheet, section_row + 3, 2)),
            "classified_status": self.reader.cell(sheet, section_row + 4, 2),
            "security_level": self.reader.cell(sheet, section_row + 5, 2),
            "s_value": self._int_or_empty(self.reader.cell(sheet, section_row + 6, 2)),
            "a_value": self._int_or_empty(self.reader.cell(sheet, section_row + 7, 2)),
            "level_consistency": self.reader.cell(sheet, section_row + 8, 2),
            "assessment_status": self.reader.cell(sheet, section_row + 9, 2),
            "assessment_org": self.reader.cell(sheet, section_row + 10, 2),
            "assessment_date": self._date_text(self.reader.cell(sheet, section_row + 11, 2)),
            "assessment_result": self.reader.cell(sheet, section_row + 12, 2),
            "pass_rate": self.reader.cell(sheet, section_row + 13, 2),
            "upstream_systems": self.reader.cell(sheet, section_row + 14, 2),
            "downstream_systems": self.reader.cell(sheet, section_row + 15, 2),
            "other_systems": self.reader.cell(sheet, section_row + 16, 2),
        }
        self.basic_repo.upsert_system_info(session, project_id, system_values)
        names = self._split_subsystems(system_values["subsystem_description"])
        self.basic_repo.sync_subsystems(session, project_id, names)
        crypto_row = self._find_row(sheet, "三、密码应用情况")
        if crypto_row is None:
            return
        crypto_values = {
            "last_assessment_status": self.reader.cell(sheet, crypto_row + 1, 2),
            "last_assessment_org": self.reader.cell(sheet, crypto_row + 2, 2),
            "last_assessment_date": self._date_text(self.reader.cell(sheet, crypto_row + 3, 2)),
            "last_assessment_result": self.reader.cell(sheet, crypto_row + 4, 2),
            "last_assessment_score": self.reader.cell(sheet, crypto_row + 5, 2),
            "has_scheme": self.reader.cell(sheet, crypto_row + 6, 2),
            "is_reviewed": self.reader.cell(sheet, crypto_row + 7, 2),
            "review_method": self.reader.cell(sheet, crypto_row + 8, 2),
            "review_date": self._date_text(self.reader.cell(sheet, crypto_row + 9, 2)),
        }
        self.basic_repo.upsert_crypto_application_info(session, project_id, crypto_values)

    def import_physical(
        self,
        session: Session,
        project_id: int,
        sheet,
        *,
        replace: bool,
    ) -> None:
        if replace:
            self._delete_physical(session, project_id)
        for row_index in range(2, sheet.max_row + 1):
            object_values = self._row_segment(sheet, row_index, 1, PHYSICAL_OBJECT_COLUMNS)
            if not object_values["name"]:
                continue
            obj = self.physical_repo.create_object(session, project_id, str(object_values["name"]))
            obj.interview_record = str(object_values["interview_record"] or "")
            obj.location = str(object_values["location"] or "")
            obj.access_control_system = str(object_values["access_control_system"] or "")
            obj.video_system = str(object_values["video_system"] or "")
            auth = self.physical_repo.load_auth_detail(session, obj.id)
            access = self.physical_repo.load_access_integrity_detail(session, obj.id)
            video = self.physical_repo.load_video_integrity_detail(session, obj.id)
            assert auth is not None and access is not None and video is not None
            auth_start = 1 + len(PHYSICAL_OBJECT_COLUMNS)
            access_start = auth_start + len(PHYSICAL_AUTH_COLUMNS)
            video_start = access_start + len(PHYSICAL_INTEGRITY_COLUMNS)
            auth_values = self._row_segment(sheet, row_index, auth_start, PHYSICAL_AUTH_COLUMNS)
            access_values = self._row_segment(
                sheet,
                row_index,
                access_start,
                PHYSICAL_INTEGRITY_COLUMNS,
            )
            video_values = self._row_segment(
                sheet,
                row_index,
                video_start,
                PHYSICAL_INTEGRITY_COLUMNS,
            )
            self._apply_detail(auth, auth_values, specific=("auth_methods",))
            self._apply_extra(
                auth,
                auth_values,
                (
                    "guarded",
                    "registered",
                    "accompanied",
                    "realtime_monitoring",
                    "mitigation_available",
                    "mitigation_note",
                    "mitigated_level",
                ),
            )
            self._save_quant_from_values(
                session,
                project_id,
                PHYSICAL_UNITS["auth"]["unit_type"],
                auth.id,
                auth_values,
                ImportCellContext("物理", SHEET_NAMES.physical, row_index),
                d_column=self._quant_d_column(auth_start, PHYSICAL_AUTH_COLUMNS),
            )
            self._save_products_from_text(
                session,
                project_id,
                PHYSICAL_UNITS["auth"]["unit_type"],
                obj.id,
                auth_values.get("products"),
            )
            for unit_key, detail, values in (
                ("access_integrity", access, access_values),
                ("video_integrity", video, video_values),
            ):
                self._apply_detail(detail, values, specific=("implementation", "algorithm"))
                self._save_quant_from_values(
                    session,
                    project_id,
                    PHYSICAL_UNITS[unit_key]["unit_type"],
                    detail.id,
                    values,
                    ImportCellContext("物理", SHEET_NAMES.physical, row_index),
                    d_column=self._quant_d_column(
                        access_start if unit_key == "access_integrity" else video_start,
                        PHYSICAL_INTEGRITY_COLUMNS,
                    ),
                )
                self._save_products_from_text(
                    session,
                    project_id,
                    PHYSICAL_UNITS[unit_key]["unit_type"],
                    obj.id,
                    values.get("products"),
                )

    def import_device(
        self,
        session: Session,
        project_id: int,
        sheet,
        *,
        replace: bool,
    ) -> None:
        if replace:
            self._delete_device(session, project_id)
        for row_index in range(2, sheet.max_row + 1):
            object_values = self._row_segment(sheet, row_index, 1, DEVICE_OBJECT_COLUMNS)
            if not object_values["name"]:
                continue
            obj = self.device_repo.create_object(session, project_id, str(object_values["name"]))
            obj.interview_record = str(object_values["interview_record"] or "")
            details = {
                "auth": self.device_repo.load_auth_detail(session, obj.id),
                "remote_management": self.device_repo.load_remote_management_detail(
                    session, obj.id
                ),
                "access_integrity": self.device_repo.load_access_integrity_detail(session, obj.id),
                "log_integrity": self.device_repo.load_log_integrity_detail(session, obj.id),
                "executable_integrity": self.device_repo.load_executable_integrity_detail(
                    session,
                    obj.id,
                ),
            }
            start_column = 1 + len(DEVICE_OBJECT_COLUMNS)
            chunks = [
                ("auth", DEVICE_AUTH_COLUMNS),
                ("remote_management", DEVICE_REMOTE_COLUMNS),
                ("access_integrity", DEVICE_INTEGRITY_COLUMNS),
                ("log_integrity", DEVICE_INTEGRITY_COLUMNS),
                ("executable_integrity", DEVICE_INTEGRITY_COLUMNS),
            ]
            for unit_key, columns in chunks:
                detail = details[unit_key]
                if detail is None:
                    start_column += len(columns)
                    continue
                values = self._row_segment(sheet, row_index, start_column, columns)
                if unit_key in {
                    "access_integrity",
                    "log_integrity",
                    "executable_integrity",
                }:
                    values = self._split_device_integrity_status(values)
                self._apply_detail(
                    detail, values, specific=tuple(DEVICE_UNITS[unit_key]["specific_fields"])
                )
                self._apply_extra(detail, values, tuple(values))
                self._save_quant_from_values(
                    session,
                    project_id,
                    DEVICE_UNITS[unit_key]["unit_type"],
                    detail.id,
                    values,
                    ImportCellContext("设备", SHEET_NAMES.device, row_index),
                    d_column=self._quant_d_column(start_column, columns),
                )
                self._save_products_from_text(
                    session,
                    project_id,
                    DEVICE_UNITS[unit_key]["unit_type"],
                    obj.id,
                    values.get("products"),
                )
                start_column += len(columns)

    def import_network(
        self,
        session: Session,
        project_id: int,
        sheet,
        *,
        replace: bool,
        skip_subsystems: set[str],
    ) -> None:
        if replace:
            self._delete_network(session, project_id)
        for row_index in range(2, sheet.max_row + 1):
            object_values = self._row_segment(sheet, row_index, 1, NETWORK_OBJECT_COLUMNS)
            subsystem_name = str(object_values["subsystem_name"] or "").strip()
            channel_name = str(object_values["channel_name"] or "").strip()
            if not subsystem_name or not channel_name:
                continue
            if subsystem_name in skip_subsystems:
                continue
            subsystem = self._ensure_network_subsystem(session, project_id, subsystem_name)
            channel = self.network_repo.create_channel(session, subsystem.id, channel_name)
            channel.extra_data = {
                "interview_record": str(object_values["interview_record"] or ""),
            }
            details = {
                "auth": self.network_repo.load_auth_detail(session, channel.id),
                "integrity": self.network_repo.load_integrity_detail(session, channel.id),
                "confidentiality": self.network_repo.load_confidentiality_detail(
                    session, channel.id
                ),
                "boundary": self.network_repo.load_boundary_detail(session, channel.id),
            }
            auth_start = 1 + len(NETWORK_OBJECT_COLUMNS)
            auth_values = self._row_segment(sheet, row_index, auth_start, NETWORK_AUTH_COLUMNS)
            channel.extra_data.update(
                {
                    "network_environment": str(auth_values["network_environment"] or ""),
                    "client_type": str(auth_values["client_type"] or ""),
                    "server_type": str(auth_values["server_type"] or ""),
                }
            )
            channel.protocol = str(auth_values["protocol"] or "")
            auth = details["auth"]
            assert auth is not None
            self._apply_detail(
                auth, auth_values, specific=tuple(NETWORK_UNITS["auth"]["specific_fields"])
            )
            self._apply_extra(auth, auth_values, tuple(auth_values))
            self._save_quant_from_values(
                session,
                project_id,
                NETWORK_UNITS["auth"]["unit_type"],
                auth.id,
                auth_values,
                ImportCellContext("网络", SHEET_NAMES.network, row_index),
                d_column=self._quant_d_column(auth_start, NETWORK_AUTH_COLUMNS),
            )
            self._save_products_from_text(
                session,
                project_id,
                NETWORK_UNITS["auth"]["unit_type"],
                channel.id,
                auth_values.get("products"),
            )
            start_column = auth_start + len(NETWORK_AUTH_COLUMNS)
            for unit_key in ("integrity", "confidentiality"):
                detail = details[unit_key]
                values = self._row_segment(
                    sheet, row_index, start_column, NETWORK_COMMON_UNIT_COLUMNS
                )
                if detail is not None:
                    self._apply_detail(
                        detail,
                        values,
                        specific=tuple(NETWORK_UNITS[unit_key]["specific_fields"]),
                    )
                    self._apply_extra(detail, values, tuple(values))
                    self._save_quant_from_values(
                        session,
                        project_id,
                        NETWORK_UNITS[unit_key]["unit_type"],
                        detail.id,
                        values,
                        ImportCellContext("网络", SHEET_NAMES.network, row_index),
                        d_column=self._quant_d_column(
                            start_column,
                            NETWORK_COMMON_UNIT_COLUMNS,
                        ),
                    )
                    self._save_products_from_text(
                        session,
                        project_id,
                        NETWORK_UNITS[unit_key]["unit_type"],
                        channel.id,
                        values.get("products"),
                    )
                start_column += len(NETWORK_COMMON_UNIT_COLUMNS)
            boundary = details["boundary"]
            boundary_values = self._row_segment(
                sheet,
                row_index,
                start_column,
                NETWORK_BOUNDARY_COLUMNS,
            )
            if boundary is not None:
                self._apply_detail(
                    boundary,
                    boundary_values,
                    specific=tuple(NETWORK_UNITS["boundary"]["specific_fields"]),
                )
                self._apply_extra(boundary, boundary_values, tuple(boundary_values))
                self._save_quant_from_values(
                    session,
                    project_id,
                    NETWORK_UNITS["boundary"]["unit_type"],
                    boundary.id,
                    boundary_values,
                    ImportCellContext("网络", SHEET_NAMES.network, row_index),
                    d_column=self._quant_d_column(start_column, NETWORK_BOUNDARY_COLUMNS),
                )

    def import_application(
        self,
        session: Session,
        project_id: int,
        sheet,
        *,
        replace: bool,
        skip_subsystems: set[str],
    ) -> None:
        if replace:
            self._delete_application(session, project_id)
        section_rows = self._application_section_rows(sheet)
        self._import_application_user_section(
            session,
            project_id,
            sheet,
            section_rows.get(APPLICATION_SECTIONS.user_auth),
            skip_subsystems,
        )
        self._import_application_access_section(
            session,
            project_id,
            sheet,
            section_rows.get(APPLICATION_SECTIONS.access_integrity),
            section_rows,
            skip_subsystems,
        )
        self._import_application_data_section(
            session,
            project_id,
            sheet,
            section_rows.get(APPLICATION_SECTIONS.important_data),
            section_rows,
            skip_subsystems,
        )
        self._import_application_business_section(
            session,
            project_id,
            sheet,
            section_rows.get(APPLICATION_SECTIONS.non_repudiation),
            section_rows,
            skip_subsystems,
        )

    def _import_application_user_section(
        self,
        session: Session,
        project_id: int,
        sheet,
        start_row: int | None,
        skip_subsystems: set[str],
    ) -> None:
        if start_row is None:
            return
        for row_index in self._section_data_rows(sheet, start_row):
            row = self.reader.row_values(sheet, row_index, APPLICATION_USER_COLUMNS)
            subsystem_name = str(row["subsystem_name"] or "").strip()
            name = str(row["name"] or "").strip()
            if not subsystem_name or not name or subsystem_name in skip_subsystems:
                continue
            subsystem = self._ensure_application_subsystem(session, project_id, subsystem_name)
            user = self.application_repo.create_user(session, subsystem.id, name)
            user.extra_data = {"interview_record": str(row["interview_record"] or "")}
            detail = self.application_repo.load_user_auth_detail(session, user.id)
            assert detail is not None
            self._apply_detail(
                detail, row, specific=tuple(APP_UNITS["user_auth"]["specific_fields"])
            )
            self._apply_extra(
                detail,
                row,
                ("key_management", "mitigation_available", "mitigation_note", "mitigated_level"),
            )
            self._save_quant_from_values(
                session,
                project_id,
                APP_UNITS["user_auth"]["unit_type"],
                detail.id,
                row,
                ImportCellContext("应用", SHEET_NAMES.application, row_index),
                d_column=self._quant_d_column(1, APPLICATION_USER_COLUMNS),
            )
            self._save_products_from_text(
                session,
                project_id,
                APP_UNITS["user_auth"]["unit_type"],
                user.id,
                row.get("products"),
            )

    def _import_application_access_section(
        self,
        session: Session,
        project_id: int,
        sheet,
        start_row: int | None,
        section_rows: dict[str, int],
        skip_subsystems: set[str],
    ) -> None:
        if start_row is None:
            return
        end_row = self._next_section_row(start_row, section_rows) or (sheet.max_row + 1)
        for row_index in range(start_row + 3, end_row):
            row = self.reader.row_values(sheet, row_index, APPLICATION_ACCESS_COLUMNS)
            subsystem_name = str(row["subsystem_name"] or "").strip()
            name = str(row["name"] or "").strip()
            if not subsystem_name or not name or subsystem_name in skip_subsystems:
                continue
            subsystem = self._ensure_application_subsystem(session, project_id, subsystem_name)
            obj = self.application_repo.create_access_control(session, subsystem.id, name)
            obj.extra_data = {"interview_record": str(row["interview_record"] or "")}
            detail = self.application_repo.load_access_integrity_detail(session, obj.id)
            if detail is None:
                continue
            self._apply_detail(
                detail, row, specific=tuple(APP_UNITS["access_integrity"]["specific_fields"])
            )
            self._apply_extra(detail, row, ("key_management",))
            self._save_quant_from_values(
                session,
                project_id,
                APP_UNITS["access_integrity"]["unit_type"],
                detail.id,
                row,
                ImportCellContext("应用", SHEET_NAMES.application, row_index),
                d_column=self._quant_d_column(1, APPLICATION_ACCESS_COLUMNS),
            )
            self._save_products_from_text(
                session,
                project_id,
                APP_UNITS["access_integrity"]["unit_type"],
                obj.id,
                row.get("products"),
            )

    def _import_application_data_section(
        self,
        session: Session,
        project_id: int,
        sheet,
        start_row: int | None,
        section_rows: dict[str, int],
        skip_subsystems: set[str],
    ) -> None:
        if start_row is None:
            return
        end_row = self._next_section_row(start_row, section_rows) or (sheet.max_row + 1)
        object_columns = ("subsystem_name", "name", "data_type", "interview_record")
        for row_index in range(start_row + 3, end_row):
            object_values = self._row_segment(sheet, row_index, 1, object_columns)
            subsystem_name = str(object_values["subsystem_name"] or "").strip()
            name = str(object_values["name"] or "").strip()
            if not subsystem_name or not name or subsystem_name in skip_subsystems:
                continue
            subsystem = self._ensure_application_subsystem(session, project_id, subsystem_name)
            data = self.application_repo.create_important_data(
                session,
                subsystem.id,
                name,
                str(object_values["data_type"] or ""),
            )
            data.extra_data = {"interview_record": str(object_values["interview_record"] or "")}
            details = {
                "transport_confidentiality": (
                    self.application_repo.load_transport_confidentiality_detail(
                        session,
                        data.id,
                    )
                ),
                "storage_confidentiality": (
                    self.application_repo.load_storage_confidentiality_detail(session, data.id)
                ),
                "transport_integrity": (
                    self.application_repo.load_transport_integrity_detail(session, data.id)
                ),
                "storage_integrity": (
                    self.application_repo.load_storage_integrity_detail(session, data.id)
                ),
            }
            start_column = 1 + len(object_columns)
            chunks = [
                ("transport_confidentiality", IMPORTANT_DATA_TRANSPORT_COLUMNS),
                ("storage_confidentiality", IMPORTANT_DATA_STORAGE_COLUMNS),
                ("transport_integrity", IMPORTANT_DATA_TRANSPORT_INTEGRITY_COLUMNS),
                ("storage_integrity", IMPORTANT_DATA_STORAGE_COLUMNS),
            ]
            for unit_key, columns in chunks:
                detail = details[unit_key]
                values = self._row_segment(sheet, row_index, start_column, columns)
                if detail is not None:
                    self._apply_application_unit_values(
                        session,
                        project_id,
                        detail,
                        unit_key,
                        values,
                    )
                    self._save_quant_from_values(
                        session,
                        project_id,
                        APP_UNITS[unit_key]["unit_type"],
                        detail.id,
                        values,
                        ImportCellContext("应用", SHEET_NAMES.application, row_index),
                        d_column=self._quant_d_column(start_column, columns),
                    )
                    self._save_products_from_text(
                        session,
                        project_id,
                        APP_UNITS[unit_key]["unit_type"],
                        data.id,
                        values.get("products"),
                    )
                start_column += len(columns)

    def _import_application_business_section(
        self,
        session: Session,
        project_id: int,
        sheet,
        start_row: int | None,
        section_rows: dict[str, int],
        skip_subsystems: set[str],
    ) -> None:
        if start_row is None:
            return
        end_row = self._next_section_row(start_row, section_rows) or (sheet.max_row + 1)
        for row_index in range(start_row + 3, end_row):
            row = self.reader.row_values(sheet, row_index, APPLICATION_BUSINESS_COLUMNS)
            subsystem_name = str(row["subsystem_name"] or "").strip()
            name = str(row["name"] or "").strip()
            if not subsystem_name or not name or subsystem_name in skip_subsystems:
                continue
            subsystem = self._ensure_application_subsystem(session, project_id, subsystem_name)
            obj = self.application_repo.create_business_action(session, subsystem.id, name)
            obj.extra_data = {"interview_record": str(row["interview_record"] or "")}
            detail = self.application_repo.load_non_repudiation_detail(session, obj.id)
            if detail is None:
                continue
            self._apply_application_unit_values(
                session,
                project_id,
                detail,
                "non_repudiation",
                row,
            )
            self._save_quant_from_values(
                session,
                project_id,
                APP_UNITS["non_repudiation"]["unit_type"],
                detail.id,
                row,
                ImportCellContext("应用", SHEET_NAMES.application, row_index),
                d_column=self._quant_d_column(1, APPLICATION_BUSINESS_COLUMNS),
            )
            self._save_products_from_text(
                session,
                project_id,
                APP_UNITS["non_repudiation"]["unit_type"],
                obj.id,
                row.get("products"),
            )

    def _apply_application_unit_values(
        self,
        session: Session,
        project_id: int,
        detail: Any,
        unit_key: str,
        values: dict[str, Any],
    ) -> None:
        self._apply_detail(detail, values, specific=tuple(APP_UNITS[unit_key]["specific_fields"]))
        if "implementation_status" in values:
            detail.implementation = str(values["implementation_status"] or "")
        if unit_key in {"transport_confidentiality", "storage_confidentiality"}:
            detail.encryption_method = str(values.get("mechanism_description") or "")
        elif unit_key in {"transport_integrity", "storage_integrity"}:
            detail.integrity_method = str(values.get("mechanism_description") or "")
        elif unit_key == "non_repudiation":
            detail.signature_method = str(values.get("mechanism_description") or "")
        if "related_channel" in values and values["related_channel"]:
            channel_id = self._find_network_channel_id(
                session, project_id, values["related_channel"]
            )
            if channel_id is not None and hasattr(detail, "network_channel_id"):
                detail.network_channel_id = channel_id
        self._apply_extra(
            detail,
            values,
            (
                "implementation_status",
                "mechanism_description",
                "image_path",
                "key_management",
                "related_channel",
                "mitigation_available",
                "mitigation_note",
                "mitigated_level",
            ),
        )

    def _save_quant_from_values(
        self,
        session: Session,
        project_id: int,
        unit_type: str,
        related_id: int,
        values: dict[str, Any],
        context: ImportCellContext,
        d_column: int | None = None,
    ) -> None:
        if d_column is not None:
            context = ImportCellContext(
                context.module,
                context.sheet,
                context.row,
                column=d_column,
            )
        quant = self.reader.parse_quant_values(
            d=values.get("d", ""),
            a=values.get("a", ""),
            k=values.get("k", ""),
            ra=values.get("ra", ""),
            rk=values.get("rk", ""),
            context=context,
        )
        self.shared_repo.delete_quant_for_related(session, project_id, unit_type, related_id)
        record = QuantitativeAssessment(
            project_id=project_id,
            unit_type=unit_type,
            related_id=related_id,
            d_value=quant["d"],
            a_value=quant["a"],
            k_value=quant["k"],
            ra_value=str(float(quant["ra"])),
            rk_value=str(float(quant["rk"])),
            score=calculate_object_score(
                d=quant["d"],
                a=quant["a"],
                k=quant["k"],
                ra=quant["ra"],
                rk=quant["rk"],
            ),
        )
        session.add(record)

    def _save_products_from_text(
        self,
        session: Session,
        project_id: int,
        unit_type: str,
        related_id: int,
        text: Any,
    ) -> None:
        self.shared_repo.delete_products_for_related(session, project_id, unit_type, related_id)
        for sort_order, product in enumerate(self.reader.parse_product_text(text)):
            self.shared_repo.add_product(
                session,
                project_id=project_id,
                unit_type=unit_type,
                related_id=related_id,
                product_name=product["name"],
                vendor=product["vendor"],
                certificate_no=product["certificate_no"],
                product_level=product["level"],
                usage=product["usage"],
                sort_order=sort_order,
            )

    def _apply_detail(
        self, detail: Any, values: dict[str, Any], *, specific: Iterable[str]
    ) -> None:
        if "implementation_status" in values:
            detail.implementation = str(values["implementation_status"] or "")
        for field in (
            "requirement",
            "implementation",
            "evaluation_result",
            "crypto_usage",
            "algorithm",
            "product_compliance",
            "compliance_status",
            "risk_level",
            "risk_analysis",
            "remediation",
            *tuple(specific),
        ):
            if field in values:
                setattr(detail, field, str(values[field] or ""))

    def _split_device_integrity_status(self, values: dict[str, Any]) -> dict[str, Any]:
        copied = dict(values)
        text = str(copied.get("implementation_status") or "").strip()
        if text.startswith(("是", "否")):
            parts = text.split(maxsplit=1)
            copied["product_used"] = parts[0]
            copied["product_level"] = parts[1] if len(parts) > 1 else ""
            copied["implementation_status"] = ""
        return copied

    def _apply_extra(self, detail: Any, values: dict[str, Any], fields: Iterable[str]) -> None:
        extra = self._mapping(getattr(detail, "extra_data", None))
        for field in fields:
            if field in values and values[field] not in (None, ""):
                extra[field] = values[field]
        if extra:
            detail.extra_data = extra

    def _ensure_basic_subsystem(self, session: Session, project_id: int, name: str):
        subsystem = session.scalar(
            select(models.Subsystem).where(
                models.Subsystem.project_id == project_id,
                models.Subsystem.name == name,
            )
        )
        if subsystem is None:
            subsystem = models.Subsystem(
                project_id=project_id,
                name=name,
                sort_order=self._next_sort(session, models.Subsystem, project_id),
            )
            session.add(subsystem)
            session.flush()
        subsystem.is_enabled = True
        return subsystem

    def _ensure_network_subsystem(
        self,
        session: Session,
        project_id: int,
        name: str,
    ) -> models.NetworkSubsystem:
        basic = self._ensure_basic_subsystem(session, project_id, name)
        subsystem = session.scalar(
            select(models.NetworkSubsystem).where(
                models.NetworkSubsystem.project_id == project_id,
                models.NetworkSubsystem.name == name,
            )
        )
        if subsystem is None:
            subsystem = models.NetworkSubsystem(
                project_id=project_id,
                basic_subsystem_id=basic.id,
                name=name,
                sort_order=basic.sort_order,
            )
            session.add(subsystem)
            session.flush()
        subsystem.basic_subsystem_id = basic.id
        subsystem.sort_order = basic.sort_order
        return subsystem

    def _ensure_application_subsystem(
        self,
        session: Session,
        project_id: int,
        name: str,
    ) -> models.ApplicationSubsystem:
        basic = self._ensure_basic_subsystem(session, project_id, name)
        subsystem = session.scalar(
            select(models.ApplicationSubsystem).where(
                models.ApplicationSubsystem.project_id == project_id,
                models.ApplicationSubsystem.name == name,
            )
        )
        if subsystem is None:
            subsystem = models.ApplicationSubsystem(
                project_id=project_id,
                basic_subsystem_id=basic.id,
                name=name,
                sort_order=basic.sort_order,
            )
            session.add(subsystem)
            session.flush()
        subsystem.basic_subsystem_id = basic.id
        subsystem.sort_order = basic.sort_order
        return subsystem

    def _find_network_channel_id(
        self,
        session: Session,
        project_id: int,
        channel_name: Any,
    ) -> int | None:
        name = str(channel_name or "").strip()
        if not name:
            return None
        return session.scalar(
            select(models.NetworkChannel.id).where(
                models.NetworkChannel.project_id == project_id,
                models.NetworkChannel.name == name,
            )
        )

    def _delete_physical(self, session: Session, project_id: int) -> None:
        for obj in self.physical_repo.list_objects(session, project_id):
            for unit_key, loader in (
                ("auth", self.physical_repo.load_auth_detail),
                ("access_integrity", self.physical_repo.load_access_integrity_detail),
                ("video_integrity", self.physical_repo.load_video_integrity_detail),
            ):
                detail = loader(session, obj.id)
                unit_type = PHYSICAL_UNITS[unit_key]["unit_type"]
                if detail is not None:
                    self.shared_repo.delete_quant_for_related(
                        session, project_id, unit_type, detail.id
                    )
                    self.shared_repo.delete_evidence_for_related(
                        session, project_id, unit_type, detail.id
                    )
                    session.delete(detail)
                self.shared_repo.delete_products_for_related(session, project_id, unit_type, obj.id)
            session.delete(obj)
        session.flush()

    def _delete_device(self, session: Session, project_id: int) -> None:
        for obj in self.device_repo.list_objects(session, project_id):
            for unit_key, loader in (
                ("auth", self.device_repo.load_auth_detail),
                ("remote_management", self.device_repo.load_remote_management_detail),
                ("access_integrity", self.device_repo.load_access_integrity_detail),
                ("log_integrity", self.device_repo.load_log_integrity_detail),
                ("executable_integrity", self.device_repo.load_executable_integrity_detail),
            ):
                detail = loader(session, obj.id)
                unit_type = DEVICE_UNITS[unit_key]["unit_type"]
                if detail is not None:
                    self.shared_repo.delete_quant_for_related(
                        session, project_id, unit_type, detail.id
                    )
                    self.shared_repo.delete_evidence_for_related(
                        session, project_id, unit_type, detail.id
                    )
                    session.delete(detail)
                self.shared_repo.delete_products_for_related(session, project_id, unit_type, obj.id)
            session.delete(obj)
        session.flush()

    def _delete_network(self, session: Session, project_id: int) -> None:
        self._clear_application_channel_references(session, project_id)
        for subsystem in self.network_repo.list_subsystems(session, project_id):
            for channel in self.network_repo.list_channels(session, subsystem.id):
                for unit_key, loader in (
                    ("auth", self.network_repo.load_auth_detail),
                    ("integrity", self.network_repo.load_integrity_detail),
                    ("confidentiality", self.network_repo.load_confidentiality_detail),
                    ("boundary", self.network_repo.load_boundary_detail),
                ):
                    detail = loader(session, channel.id)
                    unit_type = NETWORK_UNITS[unit_key]["unit_type"]
                    if detail is not None:
                        self.shared_repo.delete_quant_for_related(
                            session, project_id, unit_type, detail.id
                        )
                        self.shared_repo.delete_evidence_for_related(
                            session, project_id, unit_type, detail.id
                        )
                        self.shared_repo.delete_products_for_related(
                            session, project_id, unit_type, detail.id
                        )
                        session.delete(detail)
                    self.shared_repo.delete_products_for_related(
                        session, project_id, unit_type, channel.id
                    )
                session.delete(channel)
            session.delete(subsystem)
        session.flush()

    def _clear_application_channel_references(self, session: Session, project_id: int) -> None:
        session.execute(
            update(models.ImportantData)
            .where(models.ImportantData.project_id == project_id)
            .values(related_channel_id=None)
        )
        session.execute(
            update(models.DataTransportConfidentialityDetail)
            .where(models.DataTransportConfidentialityDetail.project_id == project_id)
            .values(network_channel_id=None)
        )
        session.execute(
            update(models.DataTransportIntegrityDetail)
            .where(models.DataTransportIntegrityDetail.project_id == project_id)
            .values(network_channel_id=None)
        )

    def _delete_application(self, session: Session, project_id: int) -> None:
        for subsystem in self.application_repo.list_subsystems(session, project_id):
            self._delete_application_users(session, project_id, subsystem.id)
            self._delete_application_access(session, project_id, subsystem.id)
            self._delete_application_data(session, project_id, subsystem.id)
            self._delete_application_business(session, project_id, subsystem.id)
            session.delete(subsystem)
        session.flush()

    def _delete_application_users(
        self, session: Session, project_id: int, subsystem_id: int
    ) -> None:
        for user in self.application_repo.list_users(session, subsystem_id):
            detail = self.application_repo.load_user_auth_detail(session, user.id)
            unit_type = APP_UNITS["user_auth"]["unit_type"]
            if detail is not None:
                self.shared_repo.delete_quant_for_related(session, project_id, unit_type, detail.id)
                self.shared_repo.delete_evidence_for_related(
                    session, project_id, unit_type, detail.id
                )
                session.delete(detail)
            self.shared_repo.delete_products_for_related(session, project_id, unit_type, user.id)
            session.delete(user)

    def _delete_application_access(
        self,
        session: Session,
        project_id: int,
        subsystem_id: int,
    ) -> None:
        for obj in self.application_repo.list_access_controls(session, subsystem_id):
            detail = self.application_repo.load_access_integrity_detail(session, obj.id)
            unit_type = APP_UNITS["access_integrity"]["unit_type"]
            if detail is not None:
                self.shared_repo.delete_quant_for_related(session, project_id, unit_type, detail.id)
                self.shared_repo.delete_evidence_for_related(
                    session, project_id, unit_type, detail.id
                )
                session.delete(detail)
            self.shared_repo.delete_products_for_related(session, project_id, unit_type, obj.id)
            session.delete(obj)

    def _delete_application_data(
        self,
        session: Session,
        project_id: int,
        subsystem_id: int,
    ) -> None:
        for data in self.application_repo.list_important_data(session, subsystem_id):
            for unit_key, loader in (
                (
                    "transport_confidentiality",
                    self.application_repo.load_transport_confidentiality_detail,
                ),
                (
                    "storage_confidentiality",
                    self.application_repo.load_storage_confidentiality_detail,
                ),
                ("transport_integrity", self.application_repo.load_transport_integrity_detail),
                ("storage_integrity", self.application_repo.load_storage_integrity_detail),
            ):
                detail = loader(session, data.id)
                unit_type = APP_UNITS[unit_key]["unit_type"]
                if detail is not None:
                    self.shared_repo.delete_quant_for_related(
                        session, project_id, unit_type, detail.id
                    )
                    self.shared_repo.delete_evidence_for_related(
                        session, project_id, unit_type, detail.id
                    )
                    session.delete(detail)
                self.shared_repo.delete_products_for_related(
                    session, project_id, unit_type, data.id
                )
            session.delete(data)

    def _delete_application_business(
        self,
        session: Session,
        project_id: int,
        subsystem_id: int,
    ) -> None:
        for obj in self.application_repo.list_business_actions(session, subsystem_id):
            detail = self.application_repo.load_non_repudiation_detail(session, obj.id)
            unit_type = APP_UNITS["non_repudiation"]["unit_type"]
            if detail is not None:
                self.shared_repo.delete_quant_for_related(session, project_id, unit_type, detail.id)
                self.shared_repo.delete_evidence_for_related(
                    session, project_id, unit_type, detail.id
                )
                session.delete(detail)
            self.shared_repo.delete_products_for_related(session, project_id, unit_type, obj.id)
            session.delete(obj)

    def _application_section_rows(self, sheet) -> dict[str, int]:
        sections = {
            APPLICATION_SECTIONS.user_auth,
            APPLICATION_SECTIONS.access_integrity,
            APPLICATION_SECTIONS.important_data,
            APPLICATION_SECTIONS.non_repudiation,
        }
        found: dict[str, int] = {}
        for row_index in range(1, sheet.max_row + 1):
            value = self.reader.cell(sheet, row_index, 1)
            if value in sections:
                found[str(value)] = row_index
        return found

    def _section_data_rows(self, sheet, start_row: int) -> range:
        end_row = sheet.max_row + 1
        for row_index in range(start_row + 1, sheet.max_row + 1):
            value = self.reader.cell(sheet, row_index, 1)
            if isinstance(value, str) and value.startswith("【"):
                end_row = row_index
                break
        return range(start_row + 3, end_row)

    def _next_section_row(self, start_row: int, section_rows: dict[str, int]) -> int | None:
        later = [row for row in section_rows.values() if row > start_row]
        return min(later) if later else None

    def _find_row(self, sheet, text: str) -> int | None:
        for row_index in range(1, sheet.max_row + 1):
            if self.reader.cell(sheet, row_index, 1) == text:
                return row_index
        return None

    def _enabled_basic_subsystem_names(self, session: Session, project_id: int) -> list[str]:
        return [str(row["name"]) for row in self.basic_repo.list_subsystems(session, project_id)]

    def _split_subsystems(self, value: Any) -> list[str]:
        return [part.strip() for part in str(value or "").split("、") if part.strip()]

    def _row_segment(
        self,
        sheet,
        row_index: int,
        start_column: int,
        columns: tuple[str, ...],
    ) -> dict[str, Any]:
        return {
            name: self.reader.cell(sheet, row_index, start_column + offset)
            for offset, name in enumerate(columns)
        }

    def _quant_d_column(self, start_column: int, columns: tuple[str, ...]) -> int:
        return start_column + columns.index("d")

    def _next_sort(self, session: Session, model, project_id: int) -> int:
        current = session.scalar(
            select(func.max(model.sort_order)).where(model.project_id == project_id)
        )
        return int(current or 0) + 1

    def _date_text(self, value: Any) -> str:
        parsed = self.reader.parse_date(value)
        return parsed.isoformat() if parsed is not None else ""

    def _int_or_empty(self, value: Any) -> int | str:
        text = str(value or "").strip()
        return int(text) if text.isdigit() else ""

    def _mapping(self, value: Any) -> dict[str, Any]:
        return dict(value) if isinstance(value, dict) else {}
