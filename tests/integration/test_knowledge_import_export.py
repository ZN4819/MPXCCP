from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from mpxccp.domain.enums import KnowledgeModule, KnowledgeType
from mpxccp.repositories.session import create_engine_for_path, init_database
from mpxccp.services.export_service import ExportService
from mpxccp.services.knowledge_service import KnowledgeService


def _engine(tmp_path: Path):
    engine = create_engine_for_path(tmp_path / "knowledge-workbook.sqlite3")
    init_database(engine)
    return engine


def _knowledge_workbook(rows: list[list[object]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "知识库"
    sheet.append(["ID", "类型", "模块", "内容", "创建时间", "更新时间"])
    for row in rows:
        sheet.append(row)
    return workbook


def test_export_knowledge_workbook_has_fixed_headers_and_records(tmp_path):
    engine = _engine(tmp_path)
    service = KnowledgeService(engine)
    added = service.add_entry(
        KnowledgeType.RISK_ANALYSIS.value,
        KnowledgeModule.PHYSICAL.value,
        "机房身份鉴别风险分析",
    )
    assert added.success

    workbook = ExportService(engine).export_knowledge_workbook()
    sheet = workbook["知识库"]

    assert workbook.sheetnames == ["知识库"]
    assert [sheet.cell(1, col).value for col in range(1, 7)] == [
        "ID",
        "类型",
        "模块",
        "内容",
        "创建时间",
        "更新时间",
    ]
    assert sheet.cell(row=2, column=2).value == KnowledgeType.RISK_ANALYSIS.value
    assert sheet.cell(row=2, column=3).value == KnowledgeModule.PHYSICAL.value
    assert sheet.cell(row=2, column=4).value == "机房身份鉴别风险分析"


def test_import_knowledge_workbook_replace_clears_old_entries(tmp_path):
    engine = _engine(tmp_path)
    service = KnowledgeService(engine)
    service.add_entry(KnowledgeType.RISK_ANALYSIS.value, KnowledgeModule.PHYSICAL.value, "旧知识")
    workbook = _knowledge_workbook(
        [
            [
                None,
                KnowledgeType.RECTIFICATION.value,
                KnowledgeModule.NETWORK.value,
                "整改建议",
                "",
                "",
            ]
        ]
    )

    result = ExportService(engine).import_knowledge_workbook(workbook, mode="替换")

    assert result.success
    assert result.payload["deleted"] == 1
    assert result.payload["added"] == 1
    assert service.list_entries(
        KnowledgeType.RISK_ANALYSIS.value,
        KnowledgeModule.PHYSICAL.value,
        show_all=True,
    ) == []
    records = service.list_entries(
        KnowledgeType.RECTIFICATION.value,
        KnowledgeModule.NETWORK.value,
        show_all=True,
    )
    assert [item.content for item in records] == ["整改建议"]


def test_import_knowledge_workbook_replace_clears_old_entries_even_when_rows_invalid(tmp_path):
    engine = _engine(tmp_path)
    service = KnowledgeService(engine)
    service.add_entry(KnowledgeType.RISK_ANALYSIS.value, KnowledgeModule.PHYSICAL.value, "旧知识")
    workbook = _knowledge_workbook([[None, "", KnowledgeModule.NETWORK.value, "", "", ""]])

    result = ExportService(engine).import_knowledge_workbook(workbook, mode="替换")

    assert result.success
    assert result.payload["deleted"] == 1
    assert result.payload["added"] == 0
    assert result.payload["skipped"] == 1
    assert service.list_all_entries() == []


def test_import_knowledge_workbook_replace_reports_skipped_invalid_rows(tmp_path):
    engine = _engine(tmp_path)
    workbook = _knowledge_workbook(
        [
            [
                None,
                KnowledgeType.RECTIFICATION.value,
                KnowledgeModule.NETWORK.value,
                "整改建议",
                "",
                "",
            ],
            [None, KnowledgeType.RECTIFICATION.value, "", "缺模块", "", ""],
        ]
    )

    result = ExportService(engine).import_knowledge_workbook(workbook, mode="替换")

    assert result.success
    assert result.payload["added"] == 1
    assert result.payload["skipped"] == 1


def test_import_knowledge_workbook_replace_rejects_missing_knowledge_sheet_without_clearing(
    tmp_path,
):
    engine = _engine(tmp_path)
    service = KnowledgeService(engine)
    service.add_entry(KnowledgeType.RISK_ANALYSIS.value, KnowledgeModule.PHYSICAL.value, "旧知识")
    workbook = Workbook()
    workbook.active.title = "错误工作表"
    workbook.active.append(["ID", "类型", "模块", "内容", "创建时间", "更新时间"])
    workbook.active.append(
        [None, KnowledgeType.RECTIFICATION.value, KnowledgeModule.NETWORK.value, "新知识", "", ""]
    )

    result = ExportService(engine).import_knowledge_workbook(workbook, mode="替换")

    assert result.success is False
    assert "知识库" in result.message
    assert [item.content for item in service.list_all_entries()] == ["旧知识"]


def test_import_knowledge_workbook_append_dedupes_by_type_module_content(tmp_path):
    engine = _engine(tmp_path)
    service = KnowledgeService(engine)
    service.add_entry(KnowledgeType.MITIGATION.value, KnowledgeModule.DEVICE.value, "已有缓释")
    workbook = _knowledge_workbook(
        [
            [
                None,
                KnowledgeType.MITIGATION.value,
                KnowledgeModule.DEVICE.value,
                "已有缓释",
                "",
                "",
            ],
            [
                None,
                KnowledgeType.MITIGATION.value,
                KnowledgeModule.DEVICE.value,
                "新增缓释",
                "",
                "",
            ],
        ]
    )

    result = ExportService(engine).import_knowledge_workbook(workbook, mode="追加")

    assert result.success
    assert result.payload["added"] == 1
    assert result.payload["skipped"] == 1
    records = service.list_entries(
        KnowledgeType.MITIGATION.value,
        KnowledgeModule.DEVICE.value,
        show_all=True,
    )
    assert [item.content for item in records] == ["已有缓释", "新增缓释"]
