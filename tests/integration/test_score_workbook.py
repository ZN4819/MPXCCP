from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from mpxccp.domain.constants import CHECK
from mpxccp.domain.enums import MeasureUnit
from mpxccp.integration.excel.score_workbook import ScoreWorkbook
from mpxccp.repositories.session import create_engine_for_path, init_database
from mpxccp.services.basic_info_service import BasicInfoService
from mpxccp.services.export_service import ExportService
from mpxccp.services.physical_service import PhysicalService
from mpxccp.services.quant_service import QuantService
from mpxccp.services.scoring_service import ScoringService

EXPECTED_SCORE_SHEETS = [
    "整体测评",
    "1物理和环境安全",
    "2网络和通信安全",
    "3设备和计算安全",
    "4应用和数据安全",
    "5管理制度",
    "6人员管理",
    "7建设运行",
    "8应急处置",
]


def _engine(tmp_path: Path):
    engine = create_engine_for_path(tmp_path / "score-workbook.sqlite3")
    init_database(engine)
    return engine


def _project_id(engine) -> int:
    result = BasicInfoService(engine).save_basic_info(
        flow_no="SCORE-XLSX-001",
        system_name="打分表测试系统",
        client_name="委托方",
        assessment_org="测评机构",
    )
    assert result.success
    assert result.project_id is not None
    return result.project_id


def test_score_workbook_has_nine_sheets_and_recalculates_summary(tmp_path):
    engine = _engine(tmp_path)
    project_id = _project_id(engine)
    scoring = ScoringService(engine)
    scoring.mark_dirty(project_id)

    workbook = ExportService(engine).export_score_workbook(project_id)
    summary = scoring.load_summary(project_id)

    assert workbook.sheetnames == EXPECTED_SCORE_SHEETS
    assert summary is not None
    assert summary.dirty is False
    overall = workbook["整体测评"]
    assert "A1:A2" in {str(item) for item in overall.merged_cells.ranges}
    assert "D1:G1" in {str(item) for item in overall.merged_cells.ranges}
    assert [overall.cell(row=2, column=column).value for column in range(4, 8)] == [
        "符合",
        "部分符合",
        "不符合",
        "不适用",
    ]


def test_score_workbook_overall_merges_layer_and_total_score_columns(tmp_path):
    engine = _engine(tmp_path)
    project_id = _project_id(engine)

    workbook = ExportService(engine).export_score_workbook(project_id)
    merged = {str(item) for item in workbook["整体测评"].merged_cells.ranges}

    assert "B3:B5" in merged
    assert "B6:B10" in merged
    assert "L3:L43" in merged


def test_technical_score_sheet_uses_real_object_name_for_effective_quant(tmp_path):
    engine = _engine(tmp_path)
    project_id = _project_id(engine)
    physical = PhysicalService(engine)
    obj = physical.create_object(project_id, "主机房")
    detail_id = physical.load_details(obj.id).auth.id
    QuantService(engine, project_id=project_id).save(
        MeasureUnit.PHYSICAL_AUTH.value,
        detail_id,
        d=CHECK,
        a=CHECK,
        k=CHECK,
        ra=1,
        rk=1,
    )

    workbook = ExportService(engine).export_score_workbook(project_id)
    sheet = workbook["1物理和环境安全"]

    assert sheet["B5"].value == "主机房"
    assert sheet["C5"].value == "适用"


def test_technical_score_sheet_marks_details_without_effective_quant_not_applicable(tmp_path):
    engine = _engine(tmp_path)
    project_id = _project_id(engine)
    PhysicalService(engine).create_object(project_id, "主机房")

    workbook = ExportService(engine).export_score_workbook(project_id)
    sheet = workbook["1物理和环境安全"]

    assert sheet["C5"].value == "不适用"
    assert str(sheet["A5"].fill.fgColor.rgb).endswith("F2F2F2")


def test_import_score_workbook_replace_clears_domain_and_parses_aliases(tmp_path):
    engine = _engine(tmp_path)
    project_id = _project_id(engine)
    scoring = ScoringService(engine)
    for indicator_no in range(23, 29):
        scoring.save_management_score(project_id, indicator_no, "不符合")
    workbook = _management_workbook({"5管理制度": ["√", 0.5, "×", "/", "", ""]})

    result = ExportService(engine).import_score_workbook(project_id, workbook, mode="替换")
    summary = scoring.load_summary(project_id)
    assert summary is not None
    details = {detail.indicator_no: detail for detail in summary.details}

    assert result.success is True
    assert any("6人员管理" in warning for warning in result.warnings)
    assert details[23].compliance_status == "符合"
    assert details[23].score == 1.0
    assert details[24].compliance_status == "部分符合"
    assert details[24].score == 0.5
    assert details[25].compliance_status == "不符合"
    assert details[25].score == 0.0
    assert details[26].compliance_status == "不适用"
    assert details[26].score is None
    assert details[27].score is None
    assert details[28].score is None


def test_import_score_workbook_merge_keeps_blank_and_unparseable_values(tmp_path):
    engine = _engine(tmp_path)
    project_id = _project_id(engine)
    scoring = ScoringService(engine)
    scoring.save_management_score(project_id, 29, "符合")
    scoring.save_management_score(project_id, 30, "不符合")
    workbook = _management_workbook({"6人员管理": [0.5, "无法解析"]})

    result = ExportService(engine).import_score_workbook(project_id, workbook, mode="合并")
    summary = scoring.load_summary(project_id)
    assert summary is not None
    details = {detail.indicator_no: detail for detail in summary.details}

    assert result.success is True
    assert details[29].compliance_status == "部分符合"
    assert details[29].score == 0.5
    assert details[30].compliance_status == "不符合"
    assert details[30].score == 0.0


def test_import_score_workbook_rolls_back_when_summary_refresh_fails(tmp_path):
    engine = _engine(tmp_path)
    project_id = _project_id(engine)
    scoring = ScoringService(engine)
    scoring.save_management_score(project_id, 23, "符合")
    workbook = _management_workbook({"5管理制度": ["不符合"]})

    result = ScoreWorkbook(
        engine,
        scoring_service=_FailingScoringService(),
    ).import_management_scores(project_id, workbook, mode="替换")
    summary = scoring.load_summary(project_id)
    assert summary is not None
    details = {detail.indicator_no: detail for detail in summary.details}

    assert result.success is False
    assert "汇总刷新失败" in result.message
    assert details[23].compliance_status == "符合"
    assert details[23].score == 1.0


def _management_workbook(sheet_values: dict[str, list[object]]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, values in sheet_values.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.cell(row=1, column=1, value="序号")
        sheet.cell(row=2, column=2, value="测评对象")
        sheet.cell(row=3, column=1, value=1)
        sheet.cell(row=3, column=2, value="管理体系")
        for offset, value in enumerate(values, start=3):
            sheet.cell(row=3, column=offset, value=value)
    return workbook


class _FailingScoringService:
    def calculate_and_persist_summary_in_session(self, *args, **kwargs):
        raise RuntimeError("汇总刷新失败")
