from __future__ import annotations

from openpyxl import Workbook

from mpxccp.domain.constants import CHECK, CROSS, SLASH
from mpxccp.integration.excel.schema import (
    APPLICATION_SECTIONS,
    DEVICE_AUTH_COLUMNS,
    DEVICE_INTEGRITY_COLUMNS,
    DEVICE_REMOTE_COLUMNS,
    SHEET_NAMES,
)


class WorkbookBuilders:
    def physical_workbook(self) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAMES.physical
        sheet.append(["测评对象名称"])
        sheet.append(self._physical_row("主机房", "机房身份鉴别模块", ra=1.2))
        return workbook

    def physical_workbook_with_invalid_ra(self) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAMES.physical
        sheet.append(["测评对象名称"])
        sheet.append(self._physical_row("主机房A", "机房A身份鉴别模块", ra=1.0))
        sheet.append(self._physical_row("主机房B", "机房B身份鉴别模块", ra="not-a-number"))
        return workbook

    def network_application_with_existing_and_new_subsystems(self) -> Workbook:
        workbook = Workbook()
        network = workbook.active
        network.title = SHEET_NAMES.network
        network.append(["子系统名称", "信道名称", "现场访谈记录"])
        network.append(self._network_row("已有子系统", "旧信道"))
        network.append(self._network_row("新子系统", "新信道"))

        application = workbook.create_sheet(SHEET_NAMES.application)
        application.append([APPLICATION_SECTIONS.user_auth])
        application.append(["子系统名称", "用户名称", "现场访谈记录"])
        application.append([])
        application.append(self._application_user_row("已有子系统", "旧用户"))
        application.append(self._application_user_row("新子系统", "新用户"))
        return workbook

    def basic_info_workbook(self, *, subsystems: str = "业务系统、管理端") -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAMES.basic_info
        values = [
            "IMP-FLOW",
            "导入后的系统",
            "项目负责人",
            "13800000000",
            "评估人员",
            "13900000000",
            "2026-06-24",
        ]
        for row, value in enumerate(values, start=2):
            sheet.cell(row=row, column=2, value=value)
        sheet.cell(row=10, column=1, value="二、系统基本信息")
        system_values = [
            "系统业务功能简介",
            subsystems,
            "2025-01-01",
            "已测评",
            "三级",
            "3",
            "2",
            "一致",
            "已测评",
            "等保机构",
            "2026-01-01",
            "符合",
            "95%",
            "上游系统",
            "下游系统",
            "其他系统",
        ]
        for row, value in enumerate(system_values, start=11):
            sheet.cell(row=row, column=2, value=value)
        sheet.cell(row=28, column=1, value="三、密码应用情况")
        crypto_values = [
            "已做过密评",
            "历史测评机构",
            "2025-06-01",
            "基本符合",
            "82",
            "是",
            "是",
            "专家评审",
            "2025-12-01",
        ]
        for row, value in enumerate(crypto_values, start=29):
            sheet.cell(row=row, column=2, value=value)
        return workbook

    def application_workbook_with_data_and_business(self) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAMES.application
        sheet.append([APPLICATION_SECTIONS.important_data])
        sheet.append(["子系统名称", "重要数据名称", "数据类型", "现场访谈记录"])
        sheet.append([])
        sheet.append(self._application_important_data_row("业务系统", "交易数据"))
        sheet.append([APPLICATION_SECTIONS.non_repudiation])
        sheet.append(["子系统名称", "关键业务行为名称", "现场访谈记录"])
        sheet.append([])
        sheet.append(self._application_business_row("业务系统", "订单提交"))
        return workbook

    def device_workbook_with_split_integrity_status(self) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAMES.device
        sheet.append(["测评对象名称", "现场访谈记录"])
        row = ["服务器A", "已访谈设备管理员"]
        row.extend([""] * len(DEVICE_AUTH_COLUMNS))
        row.extend([""] * len(DEVICE_REMOTE_COLUMNS))
        row.extend(
            [
                "是 二级",
                "SM3",
                "符合",
                "符合",
                "设备访问控制产品(厂商F,证书:CERT-DEV-ACCESS,等级:二级,用途:访问控制完整性)",
                CHECK,
                CHECK,
                CHECK,
                1,
                1,
                "低风险",
                "访问控制完整性风险可控",
                "持续维护访问控制策略",
            ]
        )
        row.extend([""] * len(DEVICE_INTEGRITY_COLUMNS))
        row.extend([""] * len(DEVICE_INTEGRITY_COLUMNS))
        sheet.append(row)
        return workbook

    def network_workbook(
        self, subsystem_name: str = "新网络", channel_name: str = "新信道"
    ) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAMES.network
        sheet.append(["子系统名称", "信道名称", "现场访谈记录"])
        sheet.append(self._network_row(subsystem_name, channel_name))
        return workbook

    def _physical_row(self, object_name: str, product_name: str, *, ra: object) -> list[object]:
        row: list[object] = [
            object_name,
            "已访谈机房值守人员",
            "一层东侧",
            "门禁系统 V1",
            "视频监控系统 V2",
            "口令、门禁卡",
            "已使用",
            "SM4",
            "符合",
            "符合",
            f"{product_name}(厂商A,证书:CERT-{object_name},等级:二级,用途:身份鉴别)",
            CHECK,
            CHECK,
            CHECK,
            ra,
            1,
            "是",
            "是",
            "否",
            "是",
            "中风险",
            "身份鉴别风险可控",
            "是",
            "已落实双人值守",
            "低风险",
            "完善值守登记",
            "已实现",
            "SM3",
            "符合",
            "符合",
            f"{object_name}门禁完整性产品(厂商B,证书:CERT-ACCESS-{object_name},等级:一级,用途:门禁记录完整性)",
            CHECK,
            CHECK,
            CROSS,
            1,
            1.2,
            "低风险",
            "门禁记录完整性可核验",
            "持续保留记录",
            "已实现",
            "SM3",
            "符合",
            "符合",
            f"{object_name}视频完整性产品(厂商C,证书:CERT-VIDEO-{object_name},等级:二级,用途:视频记录完整性)",
            CROSS,
            SLASH,
            SLASH,
            1,
            1,
            "无风险",
            "",
            "",
        ]
        return row

    def _network_row(self, subsystem_name: str, channel_name: str) -> list[object]:
        return [
            subsystem_name,
            channel_name,
            f"{subsystem_name}网络访谈",
            subsystem_name,
            "互联网",
            "浏览器",
            "网关",
            "已实现",
            "SM2",
            "符合",
            "符合",
            "HTTPS",
            "SM2",
            "CA",
            "2026-01-01",
            "2027-01-01",
            "",
            f"{channel_name}网关(厂商N,证书:CERT-{channel_name},等级:二级,用途:通信身份鉴别)",
            CHECK,
            CHECK,
            CHECK,
            1,
            1,
            "低风险",
            "通信实体身份鉴别可控",
            "是",
            "已纳入证书管理",
            "低风险",
            "保持证书续期管理",
        ]

    def _application_user_row(self, subsystem_name: str, user_name: str) -> list[object]:
        return [
            subsystem_name,
            user_name,
            f"{subsystem_name}应用访谈",
            "口令、动态令牌",
            "已使用",
            "SM3",
            "符合",
            "符合",
            f"{user_name}认证模块(厂商A,证书:CERT-{user_name},等级:二级,用途:用户身份鉴别)",
            CHECK,
            CHECK,
            CHECK,
            1,
            1,
            "集中密钥管理",
            "低风险",
            "用户身份鉴别风险可控",
            "是",
            "已启用双因素",
            "低风险",
            "持续审计账号权限",
        ]

    def _application_important_data_row(self, subsystem_name: str, data_name: str) -> list[object]:
        transport_conf = self._application_data_unit(
            f"{data_name}传输机密性产品",
            d=CHECK,
            a=CHECK,
            k=CHECK,
            related_channel="业务信道",
            remediation="持续维护传输加密",
        )
        storage_conf = self._application_data_unit(
            f"{data_name}存储机密性产品",
            d=CHECK,
            a=CHECK,
            k=CROSS,
            remediation="完善存储加密",
            include_channel=False,
        )
        transport_integrity = self._application_data_unit(
            f"{data_name}传输完整性产品",
            d=CHECK,
            a=CROSS,
            k=CHECK,
            related_channel="业务信道",
            remediation="完善传输完整性校验",
            with_mitigation=False,
        )
        storage_integrity = self._application_data_unit(
            f"{data_name}存储完整性产品",
            d=CROSS,
            a=SLASH,
            k=SLASH,
            remediation="补充存储完整性保护",
            include_channel=False,
        )
        return [
            subsystem_name,
            data_name,
            "业务数据",
            f"{subsystem_name}重要数据访谈",
            *transport_conf,
            *storage_conf,
            *transport_integrity,
            *storage_integrity,
        ]

    def _application_data_unit(
        self,
        product_name: str,
        *,
        d: str,
        a: str,
        k: str,
        remediation: str,
        related_channel: str = "",
        include_channel: bool = True,
        with_mitigation: bool = True,
    ) -> list[object]:
        values: list[object] = [
            "已实现",
            "SM4",
            "已部署保护机制",
            "",
            "符合",
            "符合",
            f"{product_name}(厂商D,证书:CERT-{product_name},等级:二级,用途:{product_name})",
            d,
            a,
            k,
            1,
            1,
        ]
        if include_channel:
            values.append(related_channel)
        values.extend(["集中密钥管理", "低风险", "重要数据风险可控"])
        if with_mitigation:
            values.extend(["是", "已有缓释措施", "低风险"])
        values.append(remediation)
        return values

    def _application_business_row(self, subsystem_name: str, action_name: str) -> list[object]:
        return [
            subsystem_name,
            action_name,
            f"{subsystem_name}关键业务访谈",
            "已实现",
            "SM2",
            "已使用签名和时间戳",
            "",
            "符合",
            "符合",
            f"{action_name}签名产品(厂商E,证书:CERT-{action_name},等级:二级,用途:不可否认性)",
            CHECK,
            CHECK,
            CHECK,
            1,
            1,
            "集中密钥管理",
            "低风险",
            "关键业务行为不可否认性风险可控",
            "是",
            "已有日志和签名留存",
            "低风险",
            "持续保留签名记录",
        ]
